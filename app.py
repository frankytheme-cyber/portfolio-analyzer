from __future__ import annotations

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import re
import io
import json
import os
import textwrap
from pathlib import Path

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Portfolio Analyzer",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={},
)

# ── Design system ────────────────────────────────────────────────────────────
# Aesthetic: "Obsidian Terminal" — deep dark surfaces, luminous cyan/emerald
# accents, frosted glass panels, monospace data with humanist labels.

st.html(
    textwrap.dedent("""
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Fraunces:ital,opsz,wght,SOFT,WONK@0,9..144,300..900,0..100,0..1;1,9..144,300..900,0..100,0..1&family=Bricolage+Grotesque:opsz,wght@12..96,300..800&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
    <link href="https://fonts.googleapis.com/css2?family=Material+Symbols+Rounded:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200" rel="stylesheet">
    <style>

    /* ═══════════════════════════════════════════════════════════════
       ALMANAC — Editorial financial-paper aesthetic
       Cream stock · ink rules · Fraunces display · tabular figures
       ═══════════════════════════════════════════════════════════════ */

    .mat {
        font-family: 'Material Symbols Rounded';
        font-weight: 300;
        font-style: normal;
        font-size: 1.3rem;
        line-height: 1;
        letter-spacing: normal;
        text-transform: none;
        display: inline-block;
        white-space: nowrap;
        word-wrap: normal;
        -webkit-font-feature-settings: 'liga';
        font-feature-settings: 'liga';
        -webkit-font-smoothing: antialiased;
        font-variation-settings: 'FILL' 0, 'wght' 300, 'GRAD' 0, 'opsz' 24;
    }

    :root {
        /* Paper stock — warm cream with FT-coded blush */
        --paper:        #F4ECDF;
        --paper-warm:   #EFE5D2;
        --paper-cool:   #FAF4E8;
        --card:         #FFFFFF;
        --card-tint:    #FBF7EE;

        /* Ink */
        --ink:           #1A1815;
        --ink-soft:      #3D362D;
        --ink-secondary: #6B6258;
        --ink-muted:     #A09A8E;
        --ink-faint:     #C9C0B0;

        /* Hairline rules */
        --rule:          #1A1815;
        --rule-soft:     rgba(26,24,21,0.18);
        --rule-faint:    rgba(26,24,21,0.10);

        /* Accent — single editorial crimson */
        --accent:        #9E2A2B;
        --accent-soft:   rgba(158,42,43,0.08);

        /* Chromatic Language — paper-friendly muted */
        --cl-bond:        #1F5F5B;
        --cl-bond-tint:   rgba(31,95,91,0.07);
        --cl-etf:         #3B5A36;
        --cl-etf-tint:    rgba(59,90,54,0.07);
        --cl-cert:        #6B2E5F;
        --cl-cert-tint:   rgba(107,46,95,0.07);
        --cl-comm:        #B45309;
        --cl-comm-tint:   rgba(180,83,9,0.07);
        --cl-azione:      #1E3A8A;
        --cl-azione-tint: rgba(30,58,138,0.07);
        --cl-crypto:      #C9893A;
        --cl-crypto-tint: rgba(201,137,58,0.08);
        --cl-altro:       #6B6258;
        --cl-altro-tint:  rgba(107,98,88,0.07);

        /* Backwards-compat aliases used inline in the body */
        --accent-cyan:    var(--cl-bond);
        --accent-emerald: var(--cl-etf);
        --accent-amber:   var(--cl-comm);
        --accent-violet:  var(--cl-cert);
        --accent-blue:    var(--cl-azione);
        --accent-rose:    var(--cl-crypto);

        --text-primary:   var(--ink);
        --text-secondary: var(--ink-secondary);
        --text-muted:     var(--ink-muted);
    }

    /* ── Hide Streamlit chrome ───────────────────────────────────── */
    #MainMenu, header[data-testid="stHeader"], footer,
    .stDeployButton, [data-testid="stToolbar"],
    [data-testid="stDecoration"],
    [data-testid="collapsedControl"],
    button[data-testid="baseButton-headerNoPadding"] {
        display: none !important;
    }

    section[data-testid="stSidebar"] {
        transform: none !important;
        min-width: 300px !important;
        max-width: 340px !important;
    }

    /* ── Global paper ────────────────────────────────────────────── */
    .stApp {
        background: var(--paper) !important;
        color: var(--ink);
        font-family: 'Bricolage Grotesque', 'Helvetica Neue', sans-serif;
        font-feature-settings: 'ss01', 'ss02', 'tnum';
    }
    /* Subtle paper grain — SVG noise overlay */
    .stApp::before {
        content: '';
        position: fixed;
        inset: 0;
        background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='200' height='200'%3E%3Cfilter id='n'%3E%3CfeTurbulence baseFrequency='0.85' numOctaves='2' stitchTiles='stitch'/%3E%3CfeColorMatrix values='0 0 0 0 0.10  0 0 0 0 0.09  0 0 0 0 0.08  0 0 0 0.04 0'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
        opacity: 0.6;
        pointer-events: none;
        z-index: 0;
        mix-blend-mode: multiply;
    }
    .stApp > * { position: relative; z-index: 1; }

    /* Tabular numerals globally */
    [class*="kpi-value"], [class*="kpi-sub"],
    code, pre, .stDataFrame, .stMetric {
        font-feature-settings: 'tnum' 1, 'lnum' 1;
    }

    /* ── Sidebar — paper-tinted column with hairline rule ─────────── */
    section[data-testid="stSidebar"] {
        background: var(--paper-warm) !important;
        border-right: 1px solid var(--rule) !important;
    }
    /* Ensure inner content scrolls — Streamlit uses different test-ids
       across versions, so cover both. */
    section[data-testid="stSidebar"] [data-testid="stSidebarContent"],
    section[data-testid="stSidebar"] [data-testid="stSidebarUserContent"],
    section[data-testid="stSidebar"] > div:first-child {
        height: 100vh !important;
        overflow-y: auto !important;
        overflow-x: hidden !important;
        padding-top: 0.5rem;
    }
    section[data-testid="stSidebar"] .stMarkdown p,
    section[data-testid="stSidebar"] .stMarkdown span,
    section[data-testid="stSidebar"] label {
        color: var(--ink-secondary) !important;
    }
    section[data-testid="stSidebar"] .stFileUploader > section {
        border: 1px dashed var(--rule-soft) !important;
        border-radius: 2px !important;
        background: var(--card-tint) !important;
        transition: border-color 0.3s, background 0.3s;
    }
    section[data-testid="stSidebar"] .stFileUploader > section:hover {
        border-color: var(--accent) !important;
        background: var(--accent-soft) !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: var(--rule-soft) !important;
        margin: 1.2rem 0 !important;
    }
    section[data-testid="stSidebar"] input,
    section[data-testid="stSidebar"] textarea {
        background: var(--card) !important;
        border: 1px solid var(--rule-soft) !important;
        border-radius: 2px !important;
        color: var(--ink) !important;
        font-family: 'JetBrains Mono', monospace !important;
        font-size: 0.85rem !important;
    }
    section[data-testid="stSidebar"] input:focus,
    section[data-testid="stSidebar"] textarea:focus {
        border-color: var(--ink) !important;
        box-shadow: none !important;
    }
    section[data-testid="stSidebar"] [data-baseweb="input"] {
        background: var(--card) !important;
    }
    section[data-testid="stSidebar"] button[kind="formSubmit"],
    section[data-testid="stSidebar"] button[kind="secondary"],
    section[data-testid="stSidebar"] [data-testid="baseButton-secondaryFormSubmit"] {
        background: var(--ink) !important;
        color: var(--paper-cool) !important;
        border: 1px solid var(--ink) !important;
        border-radius: 2px !important;
        font-family: 'Bricolage Grotesque', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.82rem !important;
        letter-spacing: 0.04em !important;
        text-transform: uppercase !important;
        transition: background 0.15s !important;
    }
    section[data-testid="stSidebar"] button[kind="formSubmit"]:hover,
    section[data-testid="stSidebar"] [data-testid="baseButton-secondaryFormSubmit"]:hover {
        background: var(--accent) !important;
        border-color: var(--accent) !important;
    }
    /* Sidebar small "x" delete button */
    section[data-testid="stSidebar"] .stButton > button {
        background: transparent !important;
        color: var(--ink-muted) !important;
        border: 1px solid var(--rule-soft) !important;
        border-radius: 2px !important;
        font-family: 'JetBrains Mono', monospace !important;
        padding: 0.25rem 0.5rem !important;
    }
    section[data-testid="stSidebar"] .stButton > button:hover {
        color: var(--accent) !important;
        border-color: var(--accent) !important;
    }

    /* ── Typography ──────────────────────────────────────────────── */
    h1, h2, h3 {
        color: var(--ink) !important;
        font-family: 'Fraunces', 'Times New Roman', serif !important;
        font-variation-settings: 'opsz' 96, 'SOFT' 30, 'WONK' 0;
        font-weight: 500 !important;
        letter-spacing: -0.025em !important;
        line-height: 1.05 !important;
    }
    h4, h5, h6 {
        color: var(--ink) !important;
        font-family: 'Bricolage Grotesque', sans-serif !important;
        font-weight: 700 !important;
    }
    p, span, label, li { color: var(--ink-soft); }

    /* ── Editorial card (replaces glass-card) ────────────────────── */
    .glass-card {
        background: var(--card);
        border: 1px solid var(--ink);
        border-radius: 2px;
        padding: 1.4rem 1.6rem;
        position: relative;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
        box-shadow: 3px 3px 0 var(--rule-faint);
    }
    .glass-card:hover {
        transform: translate(-1px, -1px);
        box-shadow: 5px 5px 0 var(--rule-soft);
    }

    /* ── KPI metric cards — Chromatic Language ──────────────────── */
    .kpi-card {
        text-align: left;
        position: relative;
        padding: 1.2rem 1.4rem !important;
        border-top: 4px solid var(--ink) !important;
        border-radius: 0 !important;
    }
    .kpi-card.kpi-bond { border-top-color: var(--cl-bond) !important; }
    .kpi-card.kpi-etf  { border-top-color: var(--cl-etf) !important; }
    .kpi-card.kpi-cert { border-top-color: var(--cl-cert) !important; }

    .kpi-card .kpi-icon {
        width: 28px; height: 28px;
        border-radius: 0;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-size: 1rem;
        margin-bottom: 0.4rem;
        background: transparent !important;
        border: 1px solid currentColor;
    }
    .kpi-card .kpi-label {
        font-family: 'Bricolage Grotesque', sans-serif;
        font-size: 0.7rem;
        font-weight: 700;
        letter-spacing: 0.14em;
        text-transform: uppercase;
        color: var(--ink-muted) !important;
        margin: 0.4rem 0 0.5rem;
        padding-bottom: 0.4rem;
        border-bottom: 1px solid var(--rule-soft);
    }
    .kpi-card .kpi-value {
        font-family: 'Fraunces', serif;
        font-variation-settings: 'opsz' 144, 'SOFT' 30, 'WONK' 0;
        font-size: 1.95rem;
        font-weight: 500;
        color: var(--ink) !important;
        line-height: 1.0;
        letter-spacing: -0.02em;
        font-feature-settings: 'tnum' 1, 'lnum' 1, 'ss01' 1;
    }
    .kpi-card .kpi-sub {
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.95rem;
        font-weight: 500;
        margin-top: 0.4rem;
        letter-spacing: -0.01em;
    }

    .kpi-icon-value { color: var(--cl-bond) !important; }
    .kpi-icon-count { color: var(--cl-etf) !important; }
    .kpi-icon-top   { color: var(--cl-cert) !important; }

    /* ── Tabs — newspaper section nav ───────────────────────────── */
    div[data-baseweb="tab-list"] {
        background: transparent !important;
        border-bottom: 2px solid var(--ink) !important;
        gap: 0 !important;
        padding-left: 0 !important;
    }
    button[data-baseweb="tab"] {
        font-family: 'Bricolage Grotesque', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.78rem !important;
        letter-spacing: 0.16em !important;
        text-transform: uppercase !important;
        color: var(--ink-muted) !important;
        background: transparent !important;
        border: none !important;
        border-bottom: 3px solid transparent !important;
        padding: 0.7rem 1.4rem !important;
        margin-bottom: -2px !important;
        transition: color 0.2s, border-color 0.2s, background 0.2s !important;
    }
    button[data-baseweb="tab"]:hover {
        color: var(--ink) !important;
        background: var(--card-tint) !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: var(--accent) !important;
        border-bottom-color: var(--accent) !important;
    }
    div[data-baseweb="tab-highlight"] { display: none !important; }

    /* ── Dataframe ───────────────────────────────────────────────── */
    .stDataFrame {
        border: 1px solid var(--ink) !important;
        border-radius: 0 !important;
        overflow: hidden !important;
        background: var(--card) !important;
    }
    .stDataFrame [data-testid="stDataFrameResizable"] { background: var(--card) !important; }
    .stDataFrame thead tr th {
        background: var(--ink) !important;
        color: var(--paper-cool) !important;
        font-family: 'Bricolage Grotesque', sans-serif !important;
        font-weight: 700 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.08em !important;
        font-size: 0.72rem !important;
    }

    /* ── Buttons ─────────────────────────────────────────────────── */
    .stDownloadButton > button,
    .stButton > button {
        background: var(--ink) !important;
        border: 1px solid var(--ink) !important;
        color: var(--paper-cool) !important;
        border-radius: 2px !important;
        font-family: 'Bricolage Grotesque', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.78rem !important;
        letter-spacing: 0.12em !important;
        text-transform: uppercase !important;
        padding: 0.6rem 1.4rem !important;
        transition: all 0.18s ease !important;
        box-shadow: 3px 3px 0 var(--rule-faint) !important;
    }
    .stDownloadButton > button:hover,
    .stButton > button:hover {
        background: var(--accent) !important;
        border-color: var(--accent) !important;
        transform: translate(-1px, -1px) !important;
        box-shadow: 4px 4px 0 var(--rule-soft) !important;
    }
    /* Forza il colore del testo interno (Streamlit avvolge la label in <p>/<div>) */
    .stDownloadButton > button p,
    .stDownloadButton > button span,
    .stDownloadButton > button div,
    .stButton > button p,
    .stButton > button span,
    .stButton > button div,
    section[data-testid="stSidebar"] button[kind="formSubmit"] p,
    section[data-testid="stSidebar"] button[kind="formSubmit"] span,
    section[data-testid="stSidebar"] [data-testid="baseButton-secondaryFormSubmit"] p,
    section[data-testid="stSidebar"] [data-testid="baseButton-secondaryFormSubmit"] span {
        color: var(--paper-cool) !important;
    }
    /* Stato disabilitato — testo grigio chiaro su nero leggibile */
    .stButton > button:disabled,
    .stDownloadButton > button:disabled {
        background: var(--ink-muted) !important;
        border-color: var(--ink-muted) !important;
        cursor: not-allowed !important;
        opacity: 0.7 !important;
    }
    .stButton > button:disabled p,
    .stButton > button:disabled span,
    .stDownloadButton > button:disabled p,
    .stDownloadButton > button:disabled span {
        color: var(--paper-cool) !important;
    }

    /* ── Form inputs (main area) ─────────────────────────────────── */
    .stTextInput input, .stNumberInput input, .stTextArea textarea,
    .stSelectbox > div > div, [data-baseweb="select"] > div {
        background: var(--card) !important;
        border: 1px solid var(--rule-soft) !important;
        border-radius: 2px !important;
        color: var(--ink) !important;
        font-family: 'JetBrains Mono', monospace !important;
    }
    .stTextInput input:focus, .stNumberInput input:focus {
        border-color: var(--ink) !important;
        box-shadow: none !important;
    }

    /* Radio / slider inks */
    .stRadio label { color: var(--ink-soft) !important; }
    .stSlider [data-baseweb="slider"] [role="slider"] {
        background: var(--accent) !important;
        border-color: var(--ink) !important;
    }

    /* Metric (st.metric) */
    [data-testid="stMetricValue"] {
        font-family: 'Fraunces', serif !important;
        font-variation-settings: 'opsz' 144 !important;
        color: var(--ink) !important;
        font-weight: 500 !important;
    }
    [data-testid="stMetricLabel"] {
        font-family: 'Bricolage Grotesque', sans-serif !important;
        text-transform: uppercase !important;
        letter-spacing: 0.1em !important;
        font-size: 0.7rem !important;
        color: var(--ink-muted) !important;
    }
    [data-testid="stMetricDelta"] { font-family: 'JetBrains Mono', monospace !important; }

    /* Expander */
    .streamlit-expanderHeader, [data-testid="stExpander"] summary {
        background: var(--card) !important;
        border: 1px solid var(--rule-soft) !important;
        border-radius: 2px !important;
        font-family: 'Bricolage Grotesque', sans-serif !important;
        font-weight: 600 !important;
        color: var(--ink) !important;
    }
    [data-testid="stExpander"] {
        border: 1px solid var(--rule-soft) !important;
        border-radius: 2px !important;
        background: var(--card-tint) !important;
    }

    /* Alert blocks */
    .stAlert, [data-testid="stNotificationContentInfo"] {
        background: var(--card) !important;
        border: 1px solid var(--ink) !important;
        border-radius: 2px !important;
        color: var(--ink) !important;
    }

    /* ── Chat bubbles ────────────────────────────────────────────── */
    [data-testid="stChatMessage"] {
        border-radius: 8px !important;
        padding: 0.85rem 1.1rem 0.85rem 0.9rem !important;
        margin-bottom: 0 !important;
        border: 1px solid transparent !important;
        box-shadow: none !important;
    }

    /* User bubble — indented right, blue-tinted */
    [data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-user"]) {
        background: rgba(30,58,138,0.055) !important;
        border-color: rgba(30,58,138,0.16) !important;
        border-left: 3px solid var(--cl-azione) !important;
        margin-left: 2.5rem !important;
        margin-right: 0 !important;
    }

    /* Assistant bubble — indented left, warm paper */
    [data-testid="stChatMessage"]:has([data-testid="chatAvatarIcon-assistant"]) {
        background: var(--card) !important;
        border-color: var(--rule-soft) !important;
        border-left: 3px solid var(--accent) !important;
        margin-right: 2.5rem !important;
        margin-left: 0 !important;
    }

    /* Avatar chips */
    [data-testid="chatAvatarIcon-user"],
    [data-testid="chatAvatarIcon-assistant"] {
        width: 26px !important;
        height: 26px !important;
        min-width: 26px !important;
        border-radius: 50% !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        font-size: 0.75rem !important;
        font-weight: 700 !important;
    }
    [data-testid="chatAvatarIcon-user"] {
        background: var(--cl-azione) !important;
        color: #fff !important;
    }
    [data-testid="chatAvatarIcon-assistant"] {
        background: var(--accent) !important;
        color: #fff !important;
    }

    /* Turn separator */
    .chat-turn-sep {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        margin: 0.9rem 0 0.35rem;
        font-size: 0.68rem;
        font-family: 'JetBrains Mono', monospace;
        letter-spacing: 0.07em;
        text-transform: uppercase;
        color: var(--ink-muted);
    }
    .chat-turn-sep::before {
        content: '';
        flex: 1;
        height: 1px;
        background: var(--rule-faint);
    }

    [data-testid="stChatInput"] textarea {
        background: var(--card) !important;
        border: 1px solid rgba(26,24,21,0.35) !important;
        border-radius: 6px !important;
        color: var(--ink) !important;
        font-family: 'Bricolage Grotesque', sans-serif !important;
        font-size: 0.9rem !important;
    }
    [data-testid="stChatInput"] textarea:focus {
        border-color: var(--cl-azione) !important;
        box-shadow: 0 0 0 2px rgba(30,58,138,0.12) !important;
    }

    /* ── Hero — editorial masthead ──────────────────────────────── */
    .hero-container {
        text-align: center;
        padding: 6rem 2rem 4rem;
        position: relative;
        max-width: 720px;
        margin: 0 auto;
    }
    .hero-container::before {
        content: 'EST. MMXXVI';
        position: absolute;
        top: 3rem; left: 50%;
        transform: translateX(-50%);
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.7rem;
        letter-spacing: 0.4em;
        color: var(--ink-muted);
    }
    .hero-container::after {
        content: '';
        position: absolute;
        top: 5rem; left: 50%;
        transform: translateX(-50%);
        width: 60%;
        border-top: 1px solid var(--rule);
    }
    .hero-logo {
        font-size: 3.4rem;
        margin-bottom: 1.4rem;
        margin-top: 1.6rem;
        display: inline-block;
        color: var(--accent);
        font-weight: 700;
    }
    .hero-logo .mat {
        background: none !important;
        -webkit-text-fill-color: var(--accent) !important;
        color: var(--accent) !important;
    }
    .hero-title {
        font-family: 'Fraunces', serif !important;
        font-variation-settings: 'opsz' 144, 'SOFT' 0, 'WONK' 1;
        font-size: clamp(3.2rem, 7vw, 5rem);
        font-weight: 400;
        color: var(--ink) !important;
        margin-bottom: 0.8rem;
        letter-spacing: -0.04em;
        line-height: 0.95;
    }
    .hero-subtitle {
        font-family: 'Fraunces', serif;
        font-variation-settings: 'opsz' 14;
        font-size: 1.15rem;
        font-style: italic;
        color: var(--ink-soft) !important;
        max-width: 480px;
        margin: 1rem auto 0;
        line-height: 1.5;
    }
    .hero-hint {
        margin-top: 2.8rem;
        display: inline-flex;
        align-items: center;
        gap: 0.6rem;
        font-family: 'Bricolage Grotesque', sans-serif;
        font-size: 0.72rem;
        letter-spacing: 0.18em;
        text-transform: uppercase;
        color: var(--ink) !important;
        background: transparent;
        border: none;
        border-top: 1px solid var(--rule);
        border-bottom: 1px solid var(--rule);
        padding: 0.6rem 1.4rem;
        font-weight: 600;
    }
    .hero-hint .arrow { color: var(--accent); }

    /* ── Section label — masthead rule ──────────────────────────── */
    .section-label {
        font-family: 'Bricolage Grotesque', sans-serif;
        font-size: 0.72rem;
        font-weight: 700;
        letter-spacing: 0.22em;
        text-transform: uppercase;
        color: var(--ink) !important;
        margin: 1rem 0 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 1px solid var(--ink);
        display: flex;
        align-items: baseline;
        gap: 0.4rem;
    }
    .section-label::after {
        content: '';
        flex: 1;
        border-bottom: 1px solid var(--rule-soft);
        margin-bottom: 0.3em;
        margin-left: 0.5rem;
    }
    .section-label .mat {
        color: var(--accent) !important;
        font-variation-settings: 'FILL' 0, 'wght' 500 !important;
    }

    /* ── Class chips ─────────────────────────────────────────────── */
    .cl-chip {
        display: inline-block;
        font-family: 'JetBrains Mono', monospace;
        font-size: 0.65rem;
        font-weight: 600;
        letter-spacing: 0.08em;
        padding: 0.18rem 0.55rem;
        border-radius: 0;
        border: 1px solid currentColor;
        text-transform: uppercase;
    }
    .cl-bond  { color: var(--cl-bond);   background: var(--cl-bond-tint); }
    .cl-etf   { color: var(--cl-etf);    background: var(--cl-etf-tint); }
    .cl-cert  { color: var(--cl-cert);   background: var(--cl-cert-tint); }
    .cl-comm  { color: var(--cl-comm);   background: var(--cl-comm-tint); }
    .cl-az    { color: var(--cl-azione); background: var(--cl-azione-tint); }
    .cl-altro { color: var(--cl-altro);  background: var(--cl-altro-tint); }

    /* ── Plotly background canvas ───────────────────────────────── */
    .js-plotly-plot, .plotly {
        background: transparent !important;
    }

    /* ── Code & inline mono ─────────────────────────────────────── */
    code {
        font-family: 'JetBrains Mono', monospace !important;
        background: var(--card-tint) !important;
        border: 1px solid var(--rule-faint) !important;
        padding: 0.1rem 0.35rem !important;
        border-radius: 2px !important;
        color: var(--ink) !important;
        font-size: 0.85em !important;
    }

    /* Scrollbar */
    ::-webkit-scrollbar { width: 10px; height: 10px; }
    ::-webkit-scrollbar-track { background: var(--paper-warm); }
    ::-webkit-scrollbar-thumb {
        background: var(--ink-muted);
        border: 2px solid var(--paper-warm);
    }
    ::-webkit-scrollbar-thumb:hover { background: var(--ink); }

    /* Selection */
    ::selection { background: var(--accent); color: var(--paper-cool); }

    </style>
    """)
)


# ── Classification helpers ───────────────────────────────────────────────────

CERT_ISSUERS = [
    "vontobel", "leonteq", "marex", "bnp paribas", "bnp", "societe generale",
    "citigroup", "barclays", "mediobanca", "unicredit",
]
CERT_KEYWORDS = [
    "cash collect", "phoenix", "barrier", "memory", "step down",
    "bonus cap", "express", "reverse convertible", "outperformance",
    "autocallable",
]
ETF_KEYWORDS = [
    "ucits", "etf", "fnd", "ishares", "vanguard", "invesco",
    "lyxor", "amundi", "xtrackers", "wisdomtree", "spdr",
]
BOND_KEYWORDS = [
    "btp", "bund", "treasury", "t-note", "t-bill", "senior preferred",
    "subordinated", "zero coupon", "cct", "corp bond", "govt bond",
    "high yield corp bond", "euromts",
]

SECTOR_MAP = {
    # Technology
    "aapl": "Technology", "msft": "Technology", "nvda": "Technology",
    "goog": "Technology", "googl": "Technology", "meta": "Technology",
    "amzn": "Technology", "tsm": "Technology", "asml": "Technology",
    "stm": "Technology", "sap": "Technology", "intc": "Technology",
    "amd": "Technology", "crm": "Technology", "adbe": "Technology",
    "orcl": "Technology", "csco": "Technology", "ibm": "Technology",
    # Energy
    "eni": "Energy", "xom": "Energy", "bp": "Energy",
    "shel": "Energy", "tot": "Energy", "cvx": "Energy",
    "cop": "Energy", "eog": "Energy", "slb": "Energy",
    "tenaris": "Energy", "saipem": "Energy",
    # Finance
    "jpm": "Finance", "isp": "Finance", "ucg": "Finance",
    "intesa": "Finance", "unicredit": "Finance", "mediobanca": "Finance",
    "generali": "Finance", "bpm": "Finance", "bper": "Finance",
    "gs": "Finance", "ms": "Finance", "bnp": "Finance",
    "ing": "Finance", "hsbc": "Finance", "ubs": "Finance",
    "dbk": "Finance", "san": "Finance",
    # Healthcare
    "pfe": "Healthcare", "jnj": "Healthcare", "unh": "Healthcare",
    "mrk": "Healthcare", "abbv": "Healthcare", "lly": "Healthcare",
    "novo": "Healthcare", "azn": "Healthcare", "gsk": "Healthcare",
    "roche": "Healthcare", "sanofi": "Healthcare", "bayer": "Healthcare",
    "recordati": "Healthcare", "diasorin": "Healthcare",
    # Utilities
    "enel": "Utilities", "a2a": "Utilities", "terna": "Utilities",
    "hera": "Utilities", "snam": "Utilities", "italgas": "Utilities",
    "erg": "Utilities", "nee": "Utilities", "duke": "Utilities",
    # Consumer / Retail
    "ko": "Consumer", "pep": "Consumer", "pg": "Consumer",
    "nesn": "Consumer", "ul": "Consumer", "lvmh": "Consumer",
    "mc": "Consumer", "or": "Consumer", "campari": "Consumer",
    # Industrials
    "leonardo": "Industrials", "prysmian": "Industrials",
    "cnh": "Industrials", "stellantis": "Industrials",
    "ferrari": "Industrials", "pirelli": "Industrials",
    "hon": "Industrials", "cat": "Industrials", "ba": "Industrials",
    "siemens": "Industrials", "airbus": "Industrials",
    # Telecom
    "tim": "Telecom", "t": "Telecom", "vz": "Telecom",
    "dte": "Telecom", "orange": "Telecom",
    # Real Estate
    "spg": "Real Estate", "o": "Real Estate", "amt": "Real Estate",
}

# ── Sector keywords (per ETF tematici / settoriali) ─────────────────────────
SECTOR_KEYWORD_MAP = {
    "Technology": [
        "technology", "tech", "information tech", "digital", "semiconductor",
        "artificial intelligence", " ai ", "robotics", "automation", "cyber",
        "cloud", "software", "nasdaq", "blockchain",
    ],
    "Healthcare": [
        "healthcare", "health care", "pharma", "biotech", "medical",
        "genomic", "aging population",
    ],
    "Energy": [
        "energy", "oil", "gas", "petroleum", "clean energy", "solar",
        "wind", "renewable", "nuclear",
    ],
    "Finance": [
        "financial", "financials", "banking", "banks", "insurance",
    ],
    "Utilities": [
        "utilities", "utility", "infrastructure",
    ],
    "Real Estate": [
        "real estate", "reit", "property", "immobiliare",
    ],
    "Consumer": [
        "consumer", "retail", "staples", "discretionary", "luxury",
        "food", "beverage",
    ],
    "Industrials": [
        "industrial", "aerospace", "defense", "defence", "materials",
        "construction", "machinery",
    ],
    "Telecom": [
        "telecom", "communication", "media",
    ],
    "ESG / SRI": [
        "esg", "sri", "sustainable", "sustainability", "climate",
        "paris aligned", "social responsibility", "green",
    ],
}

# ── Geography keywords (per ETF / fondi) ────────────────────────────────────
GEOGRAPHY_KEYWORD_MAP = {
    "World": [
        "world", "all-world", "global", "msci world", "acwi",
        "all country", "ftse all",
    ],
    "USA": [
        "s&p 500", "s&p500", "usa", "us ", "u.s.", "america",
        "nasdaq", "dow jones", "russell", "msci usa",
    ],
    "Europa": [
        "europe", "euro stoxx", "europa", "msci europe", "stoxx 600",
        "eurozone", "ftse developed europe", "euro area",
    ],
    "Italia": [
        "italy", "italia", "ftse mib", "btp", "piazza affari",
    ],
    "Emergenti": [
        "emerging", "em imi", "em ", "msci em", "ftse emerging",
    ],
    "Cina": [
        "china", "cina", "msci china", "csi 300", "hang seng",
        "shanghai",
    ],
    "Giappone": [
        "japan", "giappone", "topix", "nikkei", "msci japan",
    ],
    "Asia Pacifico": [
        "asia", "pacific", "apac", "asia-pacific", "asean",
        "asia ex japan", "asia ex-japan",
    ],
    "UK": [
        "uk ", "united kingdom", "ftse 100", "ftse 250", "britain",
        "british",
    ],
    "Germania": [
        "germany", "germania", "dax",
    ],
    "America Latina": [
        "latin america", "latam", "brazil", "brasil", "mexico",
        "messico",
    ],
    "Africa / Frontiera": [
        "africa", "frontier", "frontiera",
    ],
}

# ── ISIN country prefix → geography (per azioni dirette / bond sovrani) ─────
ISIN_COUNTRY_MAP = {
    "IT": "Italia", "US": "USA", "DE": "Germania", "FR": "Europa",
    "GB": "UK", "NL": "Europa", "ES": "Europa", "CH": "Europa",
    "JP": "Giappone", "CN": "Cina", "HK": "Cina",
    "AU": "Asia Pacifico", "KR": "Asia Pacifico", "TW": "Asia Pacifico",
    "BR": "America Latina", "MX": "America Latina",
    "IE": None, "LU": None,  # domicilio fondo, non esposizione geografica
}


BOND_UNDERLYING_KEYWORDS = [
    "bond", "obbligazion", "govt", "government", "treasury", "t-bill",
    "t-note", "aggregate", "corporate", "high yield", "investment grade",
    "fixed income", "inflation linked", "euromts", "barclays",
]

MONETARY_KEYWORDS = [
    "money market", "monetario", "short maturity", "overnight",
    "cash", "liquidity", "eonia", "ester", "€str", "floating rate",
    "ultra short", "enhanced cash",
]

COMMODITY_KEYWORDS = [
    "gold", "oro", "silver", "argento", "platinum", "platino", "palladium",
    "palladio", "commodity", "commodities", "physical", "oil", "petrolio",
    "natural gas", "agriculture", "metals", "copper", "rame",
    "etc ", "etn ",
]


# Suffissi/marker tipici di azioni quotate
EQUITY_SUFFIX_KEYWORDS = [
    " inc", " inc.", " corp", " corp.", " co.", " company",
    " ltd", " ltd.", " limited", " plc",
    " s.p.a", " spa", " s.p.a.", " s.r.l", " srl",
    " ag", " se ", " se,", " se.", " a.g.",
    " n.v", " n.v.", " nv,", " nv.",
    " s.a", " s.a.", " sa,", " sa.",
    " holding", " group", " gruppo",
    " ord", " ordinary", " common stock", " adr", " gdr",
    " cl a", " cl b", " class a", " class b", " cl.a", " cl.b",
    " azione", " azioni", "azione ord",
]

# Marker per fondi tradizionali (non-ETF) — li mappiamo a ETF
FUND_KEYWORDS = [
    "sicav", "fcp", "fondo", " fund ", " funds ", "comparto",
    "blackrock global funds", "pictet -", "pictet-", "fidelity funds",
    "jpmorgan funds", "morgan stanley investment funds",
    "schroder international", "carmignac", "anima ",
]

# Liquidità / cash positions
CASH_KEYWORDS = [
    "eur cash", "usd cash", "liquidità", "liquidita",
    "saldo conto", "conto deposito", "cash account",
    "deposit account", "interest account",
]

# ISIN country prefixes che suggeriscono azioni dirette quotate
EQUITY_ISIN_PREFIXES = {
    "US", "IT", "DE", "FR", "GB", "CH", "NL", "ES", "BE", "AT",
    "PT", "FI", "DK", "SE", "NO", "JP", "HK", "CN", "AU", "CA",
    "BR", "MX", "KR", "TW", "SG", "IL",
}


def classify_asset(name: str, isin: str | None = None) -> str:
    low = name.lower()
    isin_norm = (isin or "").strip().upper()

    # 0) Cash / liquidità
    if any(k in low for k in CASH_KEYWORDS):
        return "ALTRO"

    has_issuer = any(k in low for k in CERT_ISSUERS)
    has_cert_kw = any(k in low for k in CERT_KEYWORDS)
    # 1) Certificates
    if has_issuer and has_cert_kw:
        return "CERTIFICATE"
    if has_issuer and not any(k in low for k in ETF_KEYWORDS + BOND_KEYWORDS):
        return "CERTIFICATE"

    # 2) ETF/ETC/ETN + Fondi tradizionali — smistati per sottostante
    is_etf = any(k in low for k in ETF_KEYWORDS)
    is_fund = any(k in low for k in FUND_KEYWORDS)
    # ISIN che inizia per IE/LU/FR + presenza di suffissi tipo " r " o "acc"/"dist" → fondo
    if not is_etf and not is_fund and isin_norm[:2] in ("IE", "LU"):
        if any(k in low for k in (" acc", " dist", " r ", "(acc)", "(dist)", "ucits")):
            is_fund = True

    if is_etf or is_fund:
        is_commodity = any(k in low for k in COMMODITY_KEYWORDS)
        if is_commodity:
            return "COMMODITY"
        is_monetary = any(k in low for k in MONETARY_KEYWORDS)
        if is_monetary:
            return "BOND"
        is_bond_underlying = any(k in low for k in BOND_UNDERLYING_KEYWORDS + BOND_KEYWORDS)
        if is_bond_underlying:
            return "BOND"
        return "ETF"

    # 3) Bonds diretti — keyword o pattern "anno + %"
    if any(k in low for k in BOND_KEYWORDS):
        return "BOND"
    if re.search(r"\b20\d{2}\b", low) and "%" in low:
        return "BOND"
    # ISIN che inizia per "IT000" + nome senza marker equity → probabile BTP/obbligazione
    if isin_norm.startswith("IT000") and not any(s in low for s in EQUITY_SUFFIX_KEYWORDS):
        # Se contiene cedola/maturity/scadenza tipica
        if re.search(r"\b\d{1,2}[\.,]?\d{0,3}\s*%", low) or re.search(r"\b20\d{2}\b", low):
            return "BOND"

    # 4) Azione — suffissi societari, ticker noti, ISIN azionario
    if any(k in low for k in EQUITY_SUFFIX_KEYWORDS):
        return "AZIONE"
    # Ticker matching word-boundary contro SECTOR_MAP (azioni dirette note)
    for ticker in SECTOR_MAP.keys():
        if re.search(r"\b" + re.escape(ticker) + r"\b", low):
            return "AZIONE"
    # ISIN azionario fallback
    if isin_norm[:2] in EQUITY_ISIN_PREFIXES and not isin_norm.startswith("IT000"):
        return "AZIONE"

    return "ALTRO"


def classify_subtype(name: str, classe: str) -> str:
    """Sub-classification per analisi diversificazione (financial-analyst skill)."""
    low = name.lower()

    if classe == "BOND":
        # Monetary / ultra-short — ETF monetari assimilati a obbligazionario breve
        if any(k in low for k in ("money market", "monetario", "short maturity", "overnight",
                                   "liquidity", "eonia", "ester", "€str", "enhanced cash",
                                   "ultra short", "floating rate")):
            return "Monetario / Breve Termine"
        # Government
        if any(k in low for k in ("gov", "btp", "bund", "treasury", "t-bill", "t-note", "cct", "stato")):
            if any(k in low for k in ("short", "1-3", "breve", "maturity")):
                return "Gov. Breve Termine"
            if any(k in low for k in ("25+", "long", "lungo", "2037", "2040", "2050")):
                return "Gov. Lungo Termine"
            return "Gov. Medio Termine"
        # Corporate
        if any(k in low for k in ("corporate", "corp bond", "credit")):
            if "high yield" in low:
                return "Corporate High Yield"
            return "Corporate Inv. Grade"
        if "high yield" in low:
            return "Corporate High Yield"
        # Inflation
        if any(k in low for k in ("inflation", "indicizzat")):
            return "Inflation Linked"
        # Aggregate / Global
        if any(k in low for k in ("aggregate", "global", "broad")):
            return "Obbligazionario Globale"
        if "euromts" in low or "eurozone" in low or "euro" in low:
            return "Gov. Euro"
        return "Obbligazionario Altro"

    if classe == "AZIONE":
        # Le azioni dirette vanno tutte in "Azione Singola"
        return "Azione Singola"

    if classe == "ETF":
        # Geographic
        if any(k in low for k in ("emerging", "em imi", "em ")):
            return "Azionario Emergenti"
        if any(k in low for k in ("china", "cina")):
            return "Azionario Cina"
        if any(k in low for k in ("europe", "euro stoxx", "europa", "msci europe")):
            return "Azionario Europa"
        if any(k in low for k in ("s&p 500", "usa", "us ", "america")):
            return "Azionario USA"
        if any(k in low for k in ("world", "all-world", "global", "msci world", "acwi")):
            if "small" in low:
                return "Azionario World Small Cap"
            if "value" in low:
                return "Azionario World Value"
            if "divid" in low or "div." in low or "div " in low:
                return "Azionario World Dividendo"
            return "Azionario World"
        if any(k in low for k in ("japan", "giappone", "topix", "nikkei")):
            return "Azionario Giappone"
        if "divid" in low or "div." in low:
            return "Azionario Dividendo"
        return "Azionario Altro"

    if classe == "CERTIFICATE":
        return "Certificate"

    if classe == "COMMODITY":
        if any(k in low for k in ("gold", "oro")):
            return "Oro"
        if any(k in low for k in ("silver", "argento")):
            return "Argento"
        if any(k in low for k in ("oil", "petrolio", "crude")):
            return "Petrolio"
        if any(k in low for k in ("natural gas", "gas naturale")):
            return "Gas Naturale"
        return "Commodity Altro"

    return "Altro"


def infer_sector(name: str, current: str, classe: str = "") -> str:
    """Infer sector from name using: existing value → keyword map → ticker map."""
    if pd.notna(current) and current.strip() not in ("", "N/A", "n/a", "n/d", "N/D"):
        return current.strip()
    low = name.lower()
    # Bond e Commodity: settore non applicabile nel senso equity
    if classe in ("BOND", "COMMODITY"):
        return "—"
    # 1) Keyword-based sector detection (ETF names are descriptive)
    for sector, keywords in SECTOR_KEYWORD_MAP.items():
        if any(k in low for k in keywords):
            return sector
    # 2) Ticker-based fallback — word boundary match to avoid "ms" inside "MSCI"
    for ticker, sector in SECTOR_MAP.items():
        if re.search(r"\b" + re.escape(ticker) + r"\b", low):
            return sector
    # 3) Broad/world ETF → multi-settore
    is_etf = any(k in low for k in ETF_KEYWORDS)
    if is_etf and any(k in low for k in ("world", "all-world", "global", "acwi", "ftse all", "stoxx")):
        return "Multi-settore"
    return "Non classificato"


def infer_geography(name: str, isin: str | None, classe: str) -> str:
    """Infer geographic exposure from name and ISIN."""
    low = name.lower()
    # 1) Keyword-based geography from fund/ETF name
    for geo, keywords in GEOGRAPHY_KEYWORD_MAP.items():
        if any(k in low for k in keywords):
            return geo
    # 2) ISIN prefix for direct equities and government bonds
    if pd.notna(isin) and isinstance(isin, str) and len(isin) >= 2:
        prefix = isin[:2].upper()
        # IE/LU = fund domicile, not geographic exposure — skip
        geo = ISIN_COUNTRY_MAP.get(prefix)
        if geo is not None:
            return geo
    # 3) Heuristics for bond issuers
    if classe == "BOND":
        if any(k in low for k in ("btp", "cct", "italia")):
            return "Italia"
        if "bund" in low:
            return "Germania"
        if "treasury" in low or "t-note" in low or "t-bill" in low:
            return "USA"
        if "euromts" in low or "eurozone" in low:
            return "Europa"
    # 4) Commodity = global
    if classe == "COMMODITY":
        return "Globale"
    return "Non classificato"


# ── Data cleaning ────────────────────────────────────────────────────────────

HEADER_MAP = {
    "descrizione titolo": "Nome",
    "description": "Nome",
    "nome": "Nome",
    "security name": "Nome",
    "isin": "ISIN",
    "codice isin": "ISIN",
    "isin code": "ISIN",
    "security id": "ISIN",
    "quantita": "Quantità",
    "quantità": "Quantità",
    "quantity": "Quantità",
    "qty": "Quantità",
    "prezzo acquisto": "Prezzo Acquisto",
    "prezzo_acquisto": "Prezzo Acquisto",
    "purchase price": "Prezzo Acquisto",
    "valore di mercato (eur)": "Controvalore",
    "valore di mercato": "Controvalore",
    "market value": "Controvalore",
    "controvalore": "Controvalore",
    "settore_presunto": "Settore",
    "settore": "Settore",
    "sector": "Settore",
    "data_acquisto": "Data Acquisto",
    "data acquisto": "Data Acquisto",
    "purchase date": "Data Acquisto",
    "note": "Note",
    "notes": "Note",
    # Export bancari italiani
    "descrizione": "Nome",
    "titolo": "Nome",
    "denominazione": "Nome",
    "codice isin": "ISIN",
    "cod. isin": "ISIN",
    # Controvalore — varie diciture incluse quelle con "€"
    "valore di mercato totale": "Controvalore",
    "controvalore di mercato": "Controvalore",
    "valore attuale": "Controvalore",
    "valore portafoglio": "Controvalore",
    "ctv attuale": "Controvalore",
    "ctv di mercato": "Controvalore",
    "valore di mercato €": "Controvalore",
    "valore di mercato eur": "Controvalore",
    "cambio di mercato": "Cambio Mercato",
    # Prezzi
    "p.zo di mercato": "Prezzo Mercato",
    "prezzo di mercato": "Prezzo Mercato",
    "p.zo medio di carico": "Prezzo Acquisto",
    "prezzo medio di carico": "Prezzo Acquisto",
    "prezzo medio": "Prezzo Acquisto",
    "valore di carico": "Valore Carico",
    # Quantità
    "quantità": "Quantità",
    "q.tà": "Quantità",
    "n. titoli": "Quantità",
    # Metadati
    "simbolo": "Ticker",
    "mercato": "Mercato",
    "valuta": "Valuta",
    "strumento": "Tipo Strumento",
    "cambio di carico": "Cambio Carico",
    "plus/minus": "Plus/Minus",
    "plus/minusvalenza": "Plus/Minus",
    "var%": "Var %",
    "rendimento": "Rendimento %",
    "peso": "Peso %",
    "peso %": "Peso %",
    "peso sul portafoglio": "Peso %",
    "asset class": "Classe",
    "categoria": "Settore",
    "tipo strumento": "Classe",
}


def clean_numeric(val) -> float | None:
    if pd.isna(val):
        return None
    s = str(val).strip()
    s = re.sub(r"[€$£USDEUR\s]", "", s)
    if s in ("", "N/A", "n/a", "n/d", "N/D", "-"):
        return None
    if re.match(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$", s):
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _norm_key(col: str) -> str:
    s = str(col).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    # Sort columns so "valore di mercato" variants come before less specific ones
    # This ensures Controvalore maps to the real market value, not a cambio
    def _priority(col: str) -> int:
        k = _norm_key(col)
        if "valore di mercato" in k:
            return 0
        if "controvalore" in k or "market value" in k:
            return 1
        return 2

    sorted_cols = sorted(df.columns, key=_priority)

    rename = {}
    seen: set[str] = set()
    for col in sorted_cols:
        key = _norm_key(col)
        target = HEADER_MAP.get(key)
        # Fuzzy: strip trailing currency symbols
        if target is None:
            key2 = re.sub(r"[€$£]+$", "", key).strip()
            target = HEADER_MAP.get(key2)
        if target and target not in seen:
            rename[col] = target
            seen.add(target)
    df = df.rename(columns=rename)
    # Drop any remaining duplicate columns (keep first)
    df = df.loc[:, ~df.columns.duplicated()]
    return df


def _strip_summary_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Remove summary/total rows that appear at the bottom of bank exports."""
    name_col = None
    for c in df.columns:
        if _norm_key(c) in ("titolo", "nome", "descrizione titolo", "descrizione", "denominazione"):
            name_col = c
            break
    if name_col is None:
        # Try after rename candidates
        for c in df.columns:
            if c == "Nome":
                name_col = c
                break
    if name_col is not None:
        skip_labels = {"totale", "total", "eur", "usd", "gbp", "riepilogo", "summary", ""}
        mask = df[name_col].apply(
            lambda v: str(v).strip().lower() not in skip_labels if pd.notna(v) else False
        )
        df = df[mask].reset_index(drop=True)
    else:
        # Fallback: drop fully-empty rows
        df = df.dropna(how="all").reset_index(drop=True)
    return df


def process_portfolio(df: pd.DataFrame) -> pd.DataFrame:
    df = _strip_summary_rows(df)
    df = normalize_columns(df)
    for col in ("Nome", "ISIN", "Quantità", "Controvalore"):
        if col not in df.columns:
            df[col] = None
    df["Quantità"] = df["Quantità"].apply(clean_numeric)
    df["Controvalore"] = df["Controvalore"].apply(clean_numeric)
    if "Prezzo Acquisto" in df.columns:
        df["Prezzo Acquisto"] = df["Prezzo Acquisto"].apply(clean_numeric)
    # Fallback: se Controvalore è tutto nullo, prova a calcolarlo
    if df["Controvalore"].isna().all():
        if "Valore Carico" in df.columns:
            df["Controvalore"] = df["Valore Carico"].apply(clean_numeric)
        elif "Prezzo Mercato" in df.columns and not df["Quantità"].isna().all():
            pm = df["Prezzo Mercato"].apply(clean_numeric)
            df["Controvalore"] = pm * df["Quantità"]
    df["Classe"] = df.apply(
        lambda r: classify_asset(
            str(r.get("Nome") or ""),
            str(r.get("ISIN") or "") if r.get("ISIN") is not None else None,
        ),
        axis=1,
    )
    settore_col = "Settore" if "Settore" in df.columns else None
    df["Settore"] = df.apply(
        lambda r: infer_sector(
            str(r["Nome"] or ""),
            r.get("Settore") if settore_col else None,
            r["Classe"],
        ),
        axis=1,
    )
    # Sub-classification per diversificazione
    df["Sottotipo"] = df.apply(
        lambda r: classify_subtype(str(r["Nome"] or ""), r["Classe"]),
        axis=1,
    )
    # Geographic exposure
    df["Geografia"] = df.apply(
        lambda r: infer_geography(
            str(r["Nome"] or ""),
            r.get("ISIN"),
            r["Classe"],
        ),
        axis=1,
    )
    total = df["Controvalore"].sum()
    df["Peso %"] = (df["Controvalore"] / total * 100).round(2) if total > 0 else 0
    return df


# ── Excel export ─────────────────────────────────────────────────────────────

def to_excel_styled(df: pd.DataFrame, summary: dict | None = None) -> bytes:
    """Export professionale multi-sheet seguendo xlsx skill standards.

    Parameters
    ----------
    df : pd.DataFrame – portfolio data
    summary : dict | None – optional summary metrics for Riepilogo sheet
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Portafoglio"

    # ── Style definitions (xlsx skill color coding) ──
    font_header = Font(name="Arial", bold=True, size=10, color="000000")
    font_blue = Font(name="Arial", size=10, color="0000FF")       # inputs
    font_black = Font(name="Arial", size=10, color="000000")      # formulas/computed
    font_green = Font(name="Arial", size=10, color="008000")      # cross-sheet refs
    fill_yellow = PatternFill("solid", fgColor="FFFF00")          # N/A attention
    fill_header = PatternFill("solid", fgColor="D9E2F3")
    thin = Side(style="thin")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    border_header = Border(bottom=Side(style="medium"))
    align_right = Alignment(horizontal="right")
    align_center = Alignment(horizontal="center")

    blue_cols = {"ISIN", "Nome"}
    currency_cols = {"Controvalore", "Prezzo Acquisto", "Valore Carico", "Prezzo Mercato"}
    pct_cols = {"Peso %", "Var %", "Rendimento %"}

    # ── Headers ──
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border_header
        cell.alignment = align_center

    # ── Data rows ──
    n_rows = len(df)
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(df.columns, 1):
            val = row_data[col_name]
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_all

            # Write value
            if pd.isna(val):
                cell.value = "N/D"
                cell.fill = fill_yellow
                cell.font = Font(name="Arial", size=10, color="B7791F")
                continue

            cell.value = val

            # Font color
            if col_name in blue_cols:
                cell.font = font_blue
            else:
                cell.font = font_black

            # Number formats
            if col_name in currency_cols and isinstance(val, (int, float)):
                cell.number_format = '€ #,##0.00;(€ #,##0.00);"-"'
                cell.alignment = align_right
            elif col_name in pct_cols and isinstance(val, (int, float)):
                cell.number_format = '0.0%' if val <= 1 else '0.00"%"'
                cell.alignment = align_right
            elif col_name == "Quantità" and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = align_right

    # ── Peso % as Excel formulas (xlsx skill: formulas, not hardcodes) ──
    peso_col_idx = None
    ctv_col_idx = None
    for idx, col_name in enumerate(df.columns, 1):
        if col_name == "Peso %":
            peso_col_idx = idx
        elif col_name == "Controvalore":
            ctv_col_idx = idx

    if peso_col_idx and ctv_col_idx:
        ctv_letter = get_column_letter(ctv_col_idx)
        peso_letter = get_column_letter(peso_col_idx)
        for row_idx in range(2, n_rows + 2):
            cell = ws.cell(row=row_idx, column=peso_col_idx)
            cell.value = f"={ctv_letter}{row_idx}/SUM({ctv_letter}$2:{ctv_letter}${n_rows + 1})*100"
            cell.number_format = '0.00"%"'
            cell.font = font_black
            cell.alignment = align_right
            cell.border = border_all

    # ── Totale row ──
    total_row = n_rows + 2
    ws.cell(row=total_row, column=1, value="TOTALE").font = Font(
        name="Arial", bold=True, size=10, color="000000"
    )
    ws.cell(row=total_row, column=1).border = border_all
    if ctv_col_idx:
        ctv_letter = get_column_letter(ctv_col_idx)
        total_cell = ws.cell(row=total_row, column=ctv_col_idx)
        total_cell.value = f"=SUM({ctv_letter}2:{ctv_letter}{n_rows + 1})"
        total_cell.number_format = '€ #,##0.00'
        total_cell.font = Font(name="Arial", bold=True, size=10, color="000000")
        total_cell.border = border_all
        total_cell.alignment = align_right
    if peso_col_idx:
        pct_cell = ws.cell(row=total_row, column=peso_col_idx)
        pct_cell.value = "100.00%"
        pct_cell.font = Font(name="Arial", bold=True, size=10, color="000000")
        pct_cell.border = border_all
        pct_cell.alignment = align_right

    # ── Auto-width columns ──
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row in ws.iter_rows(min_row=2, max_row=n_rows + 2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Sheet 2: Riepilogo ──
    if summary:
        ws2 = wb.create_sheet("Riepilogo")
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 20

        riepilogo_data = [
            ("Riepilogo Portafoglio", ""),
            ("", ""),
            ("Controvalore Totale", summary.get("total_value", 0)),
            ("Numero Titoli", summary.get("n_assets", 0)),
            ("Asset Più Pesante", summary.get("heaviest", "")),
            ("Peso Asset Più Pesante", summary.get("heaviest_pct", 0)),
            ("", ""),
            ("Allocazione Macro", ""),
            ("Obbligazionario %", summary.get("bond_pct", 0)),
            ("Azionario %", summary.get("eq_pct", 0)),
            ("Altro %", summary.get("other_pct", 0)),
        ]
        # Add risk metrics if available
        if "hhi" in summary:
            riepilogo_data += [
                ("", ""),
                ("Metriche di Rischio", ""),
                ("HHI (Concentrazione)", summary.get("hhi", 0)),
                ("Top-5 Concentrazione %", summary.get("top5_pct", 0)),
                ("Sottocategorie Attive", summary.get("n_subtypes", 0)),
                ("Score Diversificazione", summary.get("diversification_rating", "")),
            ]

        for row_idx, (label, value) in enumerate(riepilogo_data, 1):
            lc = ws2.cell(row=row_idx, column=1, value=label)
            vc = ws2.cell(row=row_idx, column=2, value=value)
            lc.border = border_all
            vc.border = border_all

            if row_idx in (1, 8, 13):
                lc.font = Font(name="Arial", bold=True, size=11, color="000000")
                lc.fill = fill_header
                vc.fill = fill_header
            else:
                lc.font = Font(name="Arial", size=10, color="000000")
                # Green font for cross-sheet references
                vc.font = font_green

            if isinstance(value, (int, float)):
                vc.alignment = align_right
                if "%" in label:
                    vc.number_format = '0.0"%"'
                elif "Totale" in label:
                    vc.number_format = '€ #,##0.00'
                elif "HHI" in label:
                    vc.number_format = '0.0000'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Financial-Analyst skill: portfolio report ────────────────────────────────

def generate_fa_report(df: pd.DataFrame, metrics: dict) -> dict:
    """
    Analisi completa del portafoglio ispirata alla financial-analyst skill.
    Restituisce sezioni strutturate: allocazione, concentrazione, diversificazione,
    geografia, raccomandazioni.
    """
    total = metrics.get("total_value", 0)
    n     = metrics.get("n_assets", 0)
    hhi   = metrics.get("hhi", 0)

    # ── 1. Allocazione macro ──────────────────────────────────────────────────
    class_split = df.groupby("Classe")["Controvalore"].sum().sort_values(ascending=False)
    alloc = {cls: round(v / total * 100, 1) for cls, v in class_split.items() if total > 0}

    bond_pct   = metrics.get("bond_pct", 0)
    eq_pct     = metrics.get("eq_pct", 0)
    crypto_pct = alloc.get("CRYPTO", 0)
    other_pct  = metrics.get("other_pct", 0)

    if crypto_pct > 30:
        alloc_profile = ("Speculativo", "#C9893A",
            f"Esposizione crypto molto elevata ({crypto_pct:.1f}%). "
            "Asset ad altissima volatilità e correlazione negativa nelle fasi di risk-off; "
            "adatto solo a investitori con propensione al rischio molto alta e orizzonte lungo.")
    elif crypto_pct > 15:
        alloc_profile = ("Ad alto rischio", "#C9893A",
            f"Componente crypto significativa ({crypto_pct:.1f}%) che aumenta la volatilità complessiva del portafoglio. "
            f"Restante allocazione: obbligazionario {bond_pct:.1f}%, azionario/ETF {eq_pct:.1f}%.")
    elif bond_pct > 60:
        alloc_profile = ("Difensivo", "#1F5F5B",
            f"Il portafoglio è prevalentemente obbligazionario ({bond_pct:.1f}%). "
            "Questo posizionamento riduce la volatilità ma limita il potenziale di rendimento a lungo termine.")
    elif eq_pct > 60:
        alloc_profile = ("Aggressivo", "#3B5A36",
            f"Il portafoglio è prevalentemente azionario/ETF ({eq_pct:.1f}%). "
            "Esposizione elevata alla volatilità di mercato; adatto a orizzonti temporali lunghi.")
    elif 35 <= bond_pct <= 65 and 35 <= eq_pct <= 65:
        alloc_profile = ("Bilanciato", "#6B2E5F",
            f"Allocazione equilibrata tra obbligazionario ({bond_pct:.1f}%) e azionario ({eq_pct:.1f}%). "
            "Profilo coerente con obiettivi di crescita moderata e controllo del rischio.")
    else:
        crypto_note = f", crypto {crypto_pct:.1f}%" if crypto_pct > 0 else ""
        alloc_profile = ("Misto", "#B45309",
            f"Composizione diversificata: obbligazionario {bond_pct:.1f}%, azionario {eq_pct:.1f}%"
            f"{crypto_note}, altro {other_pct:.1f}%.")

    # ── 2. Concentrazione (HHI benchmark dalla financial-analyst skill) ────────
    # HHI < 0.01 = molto diversificato, 0.01–0.15 = moderato, > 0.15 = concentrato
    if hhi < 0.06:
        hhi_label = ("Eccellente", "#3B5A36",
            f"HHI {hhi:.4f} — portafoglio ben distribuito. "
            "Il rischio è suddiviso su molte posizioni con peso equilibrato.")
    elif hhi < 0.10:
        hhi_label = ("Buono", "#3B5A36",
            f"HHI {hhi:.4f} — diversificazione soddisfacente. "
            "Qualche posizione di maggior peso, ma entro soglie accettabili.")
    elif hhi < 0.15:
        hhi_label = ("Accettabile", "#B45309",
            f"HHI {hhi:.4f} — concentrazione moderata. "
            "Alcune posizioni pesano significativamente: monitorare il rischio single-name.")
    else:
        hhi_label = ("Concentrato", "#C9893A",
            f"HHI {hhi:.4f} — portafoglio concentrato. "
            "Poche posizioni dominano il valore: alto rischio specifico.")

    # ── 3. Top holding analysis ───────────────────────────────────────────────
    df_sorted = df.dropna(subset=["Controvalore"]).sort_values("Controvalore", ascending=False)
    top5 = df_sorted.head(5)[["Nome", "Classe", "Controvalore", "Peso %"]].copy()
    top5_pct = metrics.get("top5_pct", 0)

    # ── 4. Diversificazione per sottotipo ─────────────────────────────────────
    subtype_w = df.groupby("Sottotipo")["Controvalore"].sum() / total * 100
    active_subtypes = subtype_w[subtype_w > 2].sort_values(ascending=False)
    n_sub = metrics.get("n_subtypes", 0)

    if n_sub >= 7:
        sub_comment = "Eccellente copertura per sottocategoria: il portafoglio tocca molti segmenti di mercato."
    elif n_sub >= 5:
        sub_comment = "Buona diversificazione per sottocategoria."
    elif n_sub >= 3:
        sub_comment = "Diversificazione parziale: ampliare la copertura su altri segmenti migliorerebbe il profilo di rischio."
    else:
        sub_comment = "Diversificazione per sottocategoria limitata: il portafoglio è concentrato su pochi segmenti."

    # ── 5. Analisi geografica ──────────────────────────────────────────────────
    geo_split = df.groupby("Geografia")["Controvalore"].sum().sort_values(ascending=False)
    geo_pct = {g: round(v / total * 100, 1) for g, v in geo_split.items() if total > 0}
    top_geo = list(geo_pct.items())[:3]

    non_classified_geo = geo_pct.get("Non classificato", 0)
    if non_classified_geo > 30:
        geo_note = f"Attenzione: il {non_classified_geo:.1f}% del portafoglio ha esposizione geografica non classificata."
    elif len(geo_pct) >= 4:
        geo_note = "Buona diversificazione geografica su più aree."
    else:
        geo_note = "Esposizione geografica concentrata su poche aree."

    # ── 6. Raccomandazioni operative ──────────────────────────────────────────
    recommendations = []
    if hhi >= 0.15:
        recommendations.append(("Alta priorità", "#C9893A",
            "Ridurre la concentrazione aumentando il numero di posizioni o ribilanciando le più pesanti."))
    if metrics.get("max_weight", 0) > 25:
        heaviest = df_sorted.iloc[0]["Nome"] if len(df_sorted) > 0 else "N/A"
        recommendations.append(("Alta priorità", "#C9893A",
            f'"{heaviest[:40]}" pesa il {metrics.get("max_weight",0):.1f}%: considerare un parziale alleggerimento.'))
    if bond_pct > 70:
        recommendations.append(("Moderata", "#B45309",
            "Allocazione obbligazionaria molto elevata: valutare l'aggiunta di componente azionaria per il lungo periodo."))
    if eq_pct > 80:
        recommendations.append(("Moderata", "#B45309",
            "Esposizione azionaria molto elevata: una quota obbligazionaria migliorerebbe la resilienza in fase di correzione."))
    if non_classified_geo > 30:
        recommendations.append(("Bassa priorità", "#A09A8E",
            "Verificare e completare la classificazione geografica delle posizioni non classificate."))
    if n_sub <= 2:
        recommendations.append(("Moderata", "#B45309",
            "Ampliare la diversificazione per sottocategoria aggiungendo strumenti di classi diverse."))
    if crypto_pct > 30:
        recommendations.append(("Alta priorità", "#C9893A",
            f"Esposizione crypto molto elevata ({crypto_pct:.1f}%): l'asset class è soggetta a drawdown del 70-90% in fasi di bear market. "
            "Valutare una riduzione significativa per proteggere il capitale."))
    elif crypto_pct > 15:
        recommendations.append(("Moderata", "#B45309",
            f"Le criptovalute pesano il {crypto_pct:.1f}%: monitorare attivamente la correlazione con l'equity "
            "nelle fasi di avversione al rischio e impostare livelli di uscita preventivi."))
    elif crypto_pct > 5:
        recommendations.append(("Bassa priorità", "#A09A8E",
            f"Componente crypto al {crypto_pct:.1f}%: allocazione satellite contenuta. "
            "Verificare periodicamente che non superi la soglia di tolleranza al rischio."))
    if not recommendations:
        recommendations.append(("Nessuna azione urgente", "#3B5A36",
            "Il portafoglio presenta un profilo di rischio equilibrato. Mantenere il monitoraggio periodico."))

    return {
        "alloc_profile": alloc_profile,
        "alloc": alloc,
        "hhi_label": hhi_label,
        "top5": top5,
        "top5_pct": top5_pct,
        "active_subtypes": active_subtypes,
        "sub_comment": sub_comment,
        "geo_pct": geo_pct,
        "top_geo": top_geo,
        "geo_note": geo_note,
        "recommendations": recommendations,
        "bond_pct": bond_pct,
        "eq_pct": eq_pct,
        "crypto_pct": crypto_pct,
        "other_pct": other_pct,
        "total": total,
        "n": n,
    }


# ── Portfolio risk analytics (financial-analyst skill) ──────────────────────

def compute_risk_metrics(df: pd.DataFrame) -> dict:
    """Compute portfolio-level risk & concentration metrics."""
    weights = df["Controvalore"].dropna()
    total = weights.sum()
    if total <= 0:
        return {}
    w = weights / total
    # HHI (Herfindahl-Hirschman Index): sum of squared weights
    hhi = (w ** 2).sum()
    # Top-5 concentration
    top5_pct = w.nlargest(5).sum() * 100
    # Active subtypes (with >2% weight)
    subtype_weights = df.groupby("Sottotipo")["Controvalore"].sum() / total * 100
    n_subtypes = (subtype_weights > 2).sum()
    # Asset class split
    class_split = df.groupby("Classe")["Controvalore"].sum() / total * 100
    bond_pct = class_split.get("BOND", 0)
    eq_pct = class_split.get("ETF", 0) + class_split.get("AZIONE", 0)
    crypto_pct = class_split.get("CRYPTO", 0)
    other_pct = max(0.0, 100 - bond_pct - eq_pct - crypto_pct)
    # Single-name risk
    max_weight = w.max() * 100
    # Diversification rating (based on financial-analyst benchmarks)
    if hhi < 0.06:
        div_rating = "Eccellente"
        div_color = "#3B5A36"
    elif hhi < 0.10:
        div_rating = "Buona"
        div_color = "#3B5A36"
    elif hhi < 0.15:
        div_rating = "Accettabile"
        div_color = "#B45309"
    else:
        div_rating = "Concentrato"
        div_color = "#C9893A"
    # Concentration rating
    if max_weight > 25:
        conc_rating = "Alto Rischio"
        conc_color = "#C9893A"
    elif max_weight > 15:
        conc_rating = "Moderato"
        conc_color = "#B45309"
    else:
        conc_rating = "Basso"
        conc_color = "#3B5A36"

    return {
        "hhi": round(hhi, 4),
        "top5_pct": round(top5_pct, 1),
        "n_subtypes": int(n_subtypes),
        "bond_pct": round(bond_pct, 1),
        "eq_pct": round(eq_pct, 1),
        "crypto_pct": round(crypto_pct, 1),
        "other_pct": round(other_pct, 1),
        "max_weight": round(max_weight, 1),
        "diversification_rating": div_rating,
        "diversification_color": div_color,
        "concentration_rating": conc_rating,
        "concentration_color": conc_color,
        "n_assets": len(df),
        "total_value": round(total, 2),
    }


def generate_insights(df: pd.DataFrame, metrics: dict) -> list[str]:
    """Generate natural-language portfolio insights in Italian."""
    insights = []
    n = metrics.get("n_assets", 0)
    total = metrics.get("total_value", 0)
    insights.append(
        f"Il portafoglio comprende {n} titoli per un controvalore totale di € {total:,.2f}."
    )
    # Allocation bias
    bond_pct = metrics.get("bond_pct", 0)
    eq_pct = metrics.get("eq_pct", 0)
    if bond_pct > 60:
        insights.append(
            f"L'allocazione è sbilanciata verso l'obbligazionario ({bond_pct:.1f}%), "
            "suggerendo un posizionamento difensivo."
        )
    elif eq_pct > 60:
        insights.append(
            f"L'allocazione è sbilanciata verso l'azionario ({eq_pct:.1f}%), "
            "con una maggiore esposizione alla volatilità di mercato."
        )
    elif 35 <= bond_pct <= 65 and 35 <= eq_pct <= 65:
        insights.append(
            f"L'allocazione è bilanciata tra obbligazionario ({bond_pct:.1f}%) "
            f"e azionario ({eq_pct:.1f}%)."
        )
    # Concentration
    hhi = metrics.get("hhi", 0)
    if hhi >= 0.15:
        insights.append(
            f"Attenzione: l'indice HHI ({hhi:.4f}) indica una concentrazione elevata. "
            "Si consiglia di valutare una maggiore diversificazione."
        )
    # Single-name risk
    max_w = metrics.get("max_weight", 0)
    if max_w > 20:
        heaviest = df.loc[df["Controvalore"].idxmax(), "Nome"] if not df["Controvalore"].isna().all() else "N/A"
        insights.append(
            f"Rischio single-name: \"{heaviest}\" pesa il {max_w:.1f}% del portafoglio."
        )
    # Crypto exposure
    crypto_pct = metrics.get("crypto_pct", 0)
    if crypto_pct > 25:
        insights.append(
            f"Esposizione crypto molto elevata ({crypto_pct:.1f}%): asset ad altissima volatilità, "
            "fortemente correlati nelle fasi di risk-off. Valutare una riduzione."
        )
    elif crypto_pct > 10:
        insights.append(
            f"Componente crypto al {crypto_pct:.1f}%: monitorare la correlazione con equity "
            "in fase di avversione al rischio."
        )
    elif 0 < crypto_pct <= 10:
        insights.append(
            f"Piccola allocazione crypto ({crypto_pct:.1f}%): apporto speculativo contenuto."
        )
    # Top-5 concentration
    top5 = metrics.get("top5_pct", 0)
    if top5 > 70:
        insights.append(
            f"I primi 5 titoli rappresentano il {top5:.1f}% del valore totale."
        )
    # Diversification
    n_sub = metrics.get("n_subtypes", 0)
    if n_sub >= 6:
        insights.append(
            f"Buona diversificazione per sottocategoria: {n_sub} sottotipi con peso significativo."
        )
    elif n_sub <= 3:
        insights.append(
            f"Diversificazione limitata: solo {n_sub} sottocategorie con peso superiore al 2%."
        )
    return insights


# ── Plotly theme ─────────────────────────────────────────────────────────────

# Chromatic Language: palette semantica per classe asset (Almanac — paper-friendly)
CHART_PALETTE = [
    "#1F5F5B",  # BOND       — teal
    "#3B5A36",  # ETF        — forest
    "#6B2E5F",  # CERTIFICATE— plum
    "#B45309",  # COMMODITY  — ochre
    "#1E3A8A",  # AZIONE     — navy
    "#C9893A",  # CRYPTO     — sienna
    "#9E2A2B",  # accent crimson
    "#6B6258",  # ALTRO      — stone
]

PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Bricolage Grotesque, Helvetica Neue, sans-serif", color="#3D362D", size=12),
    margin=dict(t=56, b=28, l=28, r=28),
    title_font=dict(family="Fraunces, serif", size=15, color="#1A1815"),
    legend=dict(
        font=dict(family="Bricolage Grotesque, sans-serif", size=11, color="#6B6258"),
        bgcolor="rgba(0,0,0,0)",
        borderwidth=0,
    ),
)


# ── Crypto storage (persistenza locale) ─────────────────────────────────────

CRYPTO_FILE = Path(__file__).parent / "crypto.json"


def load_crypto() -> list[dict]:
    if not CRYPTO_FILE.exists():
        return []
    try:
        with CRYPTO_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_crypto(rows: list[dict]) -> None:
    with CRYPTO_FILE.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


UPLOAD_CACHE_DIR = Path(__file__).parent / ".cache" / "uploads"
CHAT_FILE = Path(__file__).parent / ".cache" / "chat_history.json"


def save_uploads(files) -> None:
    UPLOAD_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    for old in UPLOAD_CACHE_DIR.iterdir():
        if old.is_file():
            old.unlink()
    for uf in files:
        uf.seek(0)
        (UPLOAD_CACHE_DIR / uf.name).write_bytes(uf.read())
        uf.seek(0)


def load_cached_uploads() -> list:
    if not UPLOAD_CACHE_DIR.exists():
        return []
    out = []
    for p in sorted(UPLOAD_CACHE_DIR.iterdir()):
        if not p.is_file():
            continue
        bio = io.BytesIO(p.read_bytes())
        bio.name = p.name
        out.append(bio)
    return out


def clear_uploads() -> None:
    if not UPLOAD_CACHE_DIR.exists():
        return
    for f in UPLOAD_CACHE_DIR.iterdir():
        if f.is_file():
            f.unlink()


def load_chat() -> list[dict]:
    if not CHAT_FILE.exists():
        return []
    try:
        with CHAT_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_chat(history: list[dict]) -> None:
    CHAT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with CHAT_FILE.open("w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


# ── Obiettivi (file markdown opzionale) ─────────────────────────────────────

GOALS_FILE = Path(__file__).parent / ".cache" / "goals.md"


def load_goals() -> str:
    if not GOALS_FILE.exists():
        return ""
    try:
        return GOALS_FILE.read_text(encoding="utf-8")
    except Exception:
        return ""


def save_goals(text: str) -> None:
    GOALS_FILE.parent.mkdir(parents=True, exist_ok=True)
    GOALS_FILE.write_text(text, encoding="utf-8")


def clear_goals() -> None:
    if GOALS_FILE.exists():
        GOALS_FILE.unlink()


def crypto_rows_to_df(rows: list[dict]) -> pd.DataFrame:
    """Trasforma le crypto in righe coerenti con il portafoglio (Classe=CRYPTO)."""
    if not rows:
        return pd.DataFrame()
    out = []
    for r in rows:
        qty = float(r.get("quantita") or 0)
        price = float(r.get("prezzo") or 0)
        ctv = qty * price
        nome = (r.get("nome") or "").strip()
        out.append({
            "Nome": nome,
            "ISIN": r.get("ticker") or "",
            "Quantità": qty,
            "Prezzo Acquisto": None,
            "Controvalore": ctv,
            "Classe": "CRYPTO",
            "Settore": "Crypto",
            "Sottotipo": "Crypto",
            "Geografia": "Globale",
            "Fonte": "Crypto manuale",
        })
    return pd.DataFrame(out)


# ── Liquidità manuale (persistenza locale) ──────────────────────────────────

LIQUIDITY_FILE = Path(__file__).parent / "liquidity.json"

LIQUIDITY_TYPES = [
    "Conto Corrente",
    "Conto Deposito",
    "Pronti contro Termine",
    "Cash Valuta Estera",
    "Altro",
]


def load_liquidity() -> list[dict]:
    if not LIQUIDITY_FILE.exists():
        return []
    try:
        with LIQUIDITY_FILE.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def save_liquidity(rows: list[dict]) -> None:
    with LIQUIDITY_FILE.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def liquidity_rows_to_df(rows: list[dict]) -> pd.DataFrame:
    """Trasforma la liquidità in righe del portafoglio (Classe=LIQUIDITÀ)."""
    if not rows:
        return pd.DataFrame()
    out = []
    for r in rows:
        importo = float(r.get("importo") or 0)
        nome = (r.get("nome") or "").strip()
        tipo = (r.get("tipo") or "Conto Corrente").strip()
        valuta = (r.get("valuta") or "EUR").strip().upper()
        out.append({
            "Nome": nome,
            "ISIN": "",
            "Quantità": importo,
            "Prezzo Acquisto": 1.0,
            "Controvalore": importo,
            "Classe": "LIQUIDITÀ",
            "Settore": "Liquidità",
            "Sottotipo": tipo,
            "Geografia": "—" if valuta == "EUR" else valuta,
            "Fonte": "Liquidità manuale",
        })
    return pd.DataFrame(out)


# ── Sidebar ──────────────────────────────────────────────────────────────────

# Inizializza lo stato delle crypto (caricato da disco al primo avvio)
if "crypto_rows" not in st.session_state:
    st.session_state.crypto_rows = load_crypto()
if "liquidity_rows" not in st.session_state:
    st.session_state.liquidity_rows = load_liquidity()
if "goals_text" not in st.session_state:
    st.session_state.goals_text = load_goals()
if "goals_analysis" not in st.session_state:
    st.session_state.goals_analysis = ""

with st.sidebar:
    st.markdown(
        """
        <div style="padding: 0.6rem 0 1.2rem; display:flex; align-items:baseline; gap:0.6rem;
            border-bottom: 1px solid #1A1815; margin-bottom: 1rem;">
            <span style="font-family:'Fraunces', serif; font-weight:500; font-size:1.45rem;
                color:#9E2A2B; letter-spacing:-0.04em; font-style:italic; line-height:1;
                font-variation-settings:'opsz' 144;">Almanac</span>
            <span style="font-family:'JetBrains Mono', monospace; font-size:0.62rem;
                letter-spacing:0.2em; color:#A09A8E; text-transform:uppercase;
                margin-left:auto;">Vol. I</span>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p style="font-size:0.82rem; color:#A09A8E !important; margin-bottom:1rem;">Carica uno o più file CSV o Excel per analizzare il portafoglio in modo unificato.</p>',
        unsafe_allow_html=True,
    )
    uploaded_files = st.file_uploader(
        "Trascina qui i file",
        type=["csv", "xlsx", "xls"],
        accept_multiple_files=True,
        help="CSV, Excel (.xlsx/.xls) — puoi caricare più file contemporaneamente",
    )
    if uploaded_files:
        save_uploads(uploaded_files)
    else:
        cached = load_cached_uploads()
        if cached:
            uploaded_files = cached

    if uploaded_files:
        for uf in uploaded_files:
            st.markdown(
                f'<p style="font-family:JetBrains Mono,monospace; font-size:0.72rem; color:#3B5A36 !important; margin:0.15rem 0;">'
                f'✓ {uf.name}</p>',
                unsafe_allow_html=True,
            )
        if st.button("Rimuovi file caricati", use_container_width=True, key="clear_uploads_btn"):
            clear_uploads()
            st.rerun()

    # ── Obiettivi (Markdown) ───────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        '<div style="display:flex; align-items:center; gap:0.4rem; margin-bottom:0.4rem;">'
        '<span class="mat" style="font-size:1.1rem; color:#9E2A2B; '
        "font-variation-settings:'FILL' 1,'wght' 400,'GRAD' 0,'opsz' 24;\">flag</span>"
        '<span style="font-family:Bricolage Grotesque,sans-serif; font-weight:700; font-size:0.85rem; color:#1A1815; text-transform:uppercase; letter-spacing:0.08em;">'
        "Obiettivi</span></div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p style="font-size:0.75rem; color:#A09A8E !important; margin-bottom:0.6rem;">'
        "Carica un file <code>.md</code> con i tuoi obiettivi (orizzonte, "
        "asset allocation target, vincoli di liquidità, ribilanciamento…). "
        "Verranno usati nell'analisi e nella chat.</p>",
        unsafe_allow_html=True,
    )
    goals_upload = st.file_uploader(
        "Trascina qui il file .md",
        type=["md", "markdown", "txt"],
        accept_multiple_files=False,
        key="goals_uploader",
    )
    if goals_upload is not None:
        try:
            new_text = goals_upload.read().decode("utf-8")
            if new_text != st.session_state.goals_text:
                st.session_state.goals_text = new_text
                # Reset analisi precedente perché basata su vecchi obiettivi
                st.session_state.goals_analysis = ""
                save_goals(new_text)
                st.rerun()
        except Exception as e:
            st.warning(f"Impossibile leggere il file: {e}")

    if st.session_state.goals_text:
        preview = st.session_state.goals_text.strip().splitlines()
        first_line = preview[0] if preview else "(vuoto)"
        st.markdown(
            f'<p style="font-family:JetBrains Mono,monospace; font-size:0.72rem; '
            f'color:#3B5A36 !important; margin:0.4rem 0 0.2rem;">'
            f'✓ {len(st.session_state.goals_text)} caratteri · {len(preview)} righe</p>'
            f'<p style="font-family:Fraunces,serif; font-style:italic; font-size:0.78rem; '
            f'color:#6B6258 !important; margin:0 0 0.4rem;">{first_line[:80]}{"…" if len(first_line) > 80 else ""}</p>',
            unsafe_allow_html=True,
        )
        if st.button("Rimuovi obiettivi", use_container_width=True, key="clear_goals_btn"):
            st.session_state.goals_text = ""
            st.session_state.goals_analysis = ""
            clear_goals()
            st.rerun()

    # ── Crypto manuali ─────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        '<div style="display:flex; align-items:center; gap:0.4rem; margin-bottom:0.4rem;">'
        '<span class="mat" style="font-size:1.1rem; color:#C9893A; '
        "font-variation-settings:'FILL' 1,'wght' 400,'GRAD' 0,'opsz' 24;\">currency_bitcoin</span>"
        '<span style="font-family:Bricolage Grotesque,sans-serif; font-weight:700; font-size:0.85rem; color:#1A1815; text-transform:uppercase; letter-spacing:0.08em;">'
        "Crypto manuali</span></div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p style="font-size:0.75rem; color:#A09A8E !important; margin-bottom:0.6rem;">'
        "Aggiungi posizioni in criptovalute. Salvate in <code>crypto.json</code>.</p>",
        unsafe_allow_html=True,
    )

    with st.form("crypto_add_form", clear_on_submit=True):
        c_nome = st.text_input("Nome", placeholder="Bitcoin, Ethereum, …")
        c_ticker = st.text_input("Ticker (opz.)", placeholder="BTC, ETH, …")
        c_qty = st.number_input("Quantità", min_value=0.0, value=0.0, step=0.0001, format="%.8f")
        c_price = st.number_input("Prezzo unitario (€)", min_value=0.0, value=0.0, step=1.0, format="%.2f")
        submitted = st.form_submit_button("Aggiungi", use_container_width=True)
        if submitted:
            if not c_nome.strip():
                st.warning("Inserisci un nome.")
            elif c_qty <= 0 or c_price <= 0:
                st.warning("Quantità e prezzo devono essere > 0.")
            else:
                st.session_state.crypto_rows.append({
                    "nome": c_nome.strip(),
                    "ticker": c_ticker.strip().upper(),
                    "quantita": float(c_qty),
                    "prezzo": float(c_price),
                })
                save_crypto(st.session_state.crypto_rows)
                st.rerun()

    if st.session_state.crypto_rows:
        st.markdown(
            '<p style="font-size:0.72rem; color:#6B6258 !important; margin:0.6rem 0 0.3rem; '
            'text-transform:uppercase; letter-spacing:0.06em;">In portafoglio</p>',
            unsafe_allow_html=True,
        )
        for idx, r in enumerate(list(st.session_state.crypto_rows)):
            ctv = float(r.get("quantita") or 0) * float(r.get("prezzo") or 0)
            label = (
                f"{r.get('nome','—')}"
                f"{(' · ' + r.get('ticker','')) if r.get('ticker') else ''}"
                f"  ·  € {ctv:,.2f}"
            )
            with st.expander(label, expanded=False):
                with st.form(f"crypto_edit_form_{idx}", clear_on_submit=False):
                    e_nome = st.text_input(
                        "Nome", value=str(r.get("nome", "")), key=f"e_nome_{idx}"
                    )
                    e_ticker = st.text_input(
                        "Ticker", value=str(r.get("ticker", "")), key=f"e_tk_{idx}"
                    )
                    e_qty = st.number_input(
                        "Quantità",
                        min_value=0.0,
                        value=float(r.get("quantita") or 0.0),
                        step=0.0001,
                        format="%.8f",
                        key=f"e_qty_{idx}",
                    )
                    e_price = st.number_input(
                        "Prezzo unitario (€)",
                        min_value=0.0,
                        value=float(r.get("prezzo") or 0.0),
                        step=1.0,
                        format="%.2f",
                        key=f"e_pr_{idx}",
                    )
                    bcols = st.columns([1, 1])
                    with bcols[0]:
                        save_clicked = st.form_submit_button(
                            "Salva", use_container_width=True
                        )
                    with bcols[1]:
                        del_clicked = st.form_submit_button(
                            "Rimuovi", use_container_width=True
                        )
                    if save_clicked:
                        if not e_nome.strip():
                            st.warning("Inserisci un nome.")
                        elif e_qty <= 0 or e_price <= 0:
                            st.warning("Quantità e prezzo devono essere > 0.")
                        else:
                            st.session_state.crypto_rows[idx] = {
                                "nome": e_nome.strip(),
                                "ticker": e_ticker.strip().upper(),
                                "quantita": float(e_qty),
                                "prezzo": float(e_price),
                            }
                            save_crypto(st.session_state.crypto_rows)
                            st.rerun()
                    if del_clicked:
                        st.session_state.crypto_rows.pop(idx)
                        save_crypto(st.session_state.crypto_rows)
                        st.rerun()

    # ── Liquidità manuale ──────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        '<div style="display:flex; align-items:center; gap:0.4rem; margin-bottom:0.4rem;">'
        '<span class="mat" style="font-size:1.1rem; color:#1F5F5B; '
        "font-variation-settings:'FILL' 1,'wght' 400,'GRAD' 0,'opsz' 24;\">savings</span>"
        '<span style="font-family:Bricolage Grotesque,sans-serif; font-weight:700; font-size:0.85rem; color:#1A1815; text-transform:uppercase; letter-spacing:0.08em;">'
        "Liquidità</span></div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p style="font-size:0.75rem; color:#A09A8E !important; margin-bottom:0.6rem;">'
        "Aggiungi conti correnti, depositi, cash. Salvati in <code>liquidity.json</code>.</p>",
        unsafe_allow_html=True,
    )

    with st.form("liquidity_add_form", clear_on_submit=True):
        l_nome = st.text_input("Nome", placeholder="C/C Fineco, Deposito BPM, …", key="l_add_nome")
        l_tipo = st.selectbox("Tipo", LIQUIDITY_TYPES, key="l_add_tipo")
        l_importo = st.number_input(
            "Importo (€)", min_value=0.0, value=0.0, step=100.0, format="%.2f",
            key="l_add_importo",
        )
        l_valuta = st.text_input("Valuta", value="EUR", key="l_add_valuta")
        l_submitted = st.form_submit_button("Aggiungi", use_container_width=True)
        if l_submitted:
            if not l_nome.strip():
                st.warning("Inserisci un nome.")
            elif l_importo <= 0:
                st.warning("L'importo deve essere > 0.")
            else:
                st.session_state.liquidity_rows.append({
                    "nome": l_nome.strip(),
                    "tipo": l_tipo,
                    "importo": float(l_importo),
                    "valuta": (l_valuta or "EUR").strip().upper(),
                })
                save_liquidity(st.session_state.liquidity_rows)
                st.rerun()

    if st.session_state.liquidity_rows:
        st.markdown(
            '<p style="font-size:0.72rem; color:#6B6258 !important; margin:0.6rem 0 0.3rem; '
            'text-transform:uppercase; letter-spacing:0.06em;">In portafoglio</p>',
            unsafe_allow_html=True,
        )
        for idx, r in enumerate(list(st.session_state.liquidity_rows)):
            importo = float(r.get("importo") or 0)
            valuta = (r.get("valuta") or "EUR").upper()
            label = (
                f"{r.get('nome','—')}"
                f"  ·  {r.get('tipo','')}"
                f"  ·  {valuta} {importo:,.2f}"
            )
            with st.expander(label, expanded=False):
                with st.form(f"liquidity_edit_form_{idx}", clear_on_submit=False):
                    el_nome = st.text_input(
                        "Nome", value=str(r.get("nome", "")), key=f"l_e_nome_{idx}"
                    )
                    el_tipo = st.selectbox(
                        "Tipo",
                        LIQUIDITY_TYPES,
                        index=LIQUIDITY_TYPES.index(r.get("tipo"))
                            if r.get("tipo") in LIQUIDITY_TYPES else 0,
                        key=f"l_e_tipo_{idx}",
                    )
                    el_importo = st.number_input(
                        "Importo (€)",
                        min_value=0.0,
                        value=float(r.get("importo") or 0.0),
                        step=100.0,
                        format="%.2f",
                        key=f"l_e_imp_{idx}",
                    )
                    el_valuta = st.text_input(
                        "Valuta",
                        value=str(r.get("valuta", "EUR")),
                        key=f"l_e_val_{idx}",
                    )
                    lcols = st.columns([1, 1])
                    with lcols[0]:
                        l_save_clicked = st.form_submit_button(
                            "Salva", use_container_width=True
                        )
                    with lcols[1]:
                        l_del_clicked = st.form_submit_button(
                            "Rimuovi", use_container_width=True
                        )
                    if l_save_clicked:
                        if not el_nome.strip():
                            st.warning("Inserisci un nome.")
                        elif el_importo <= 0:
                            st.warning("L'importo deve essere > 0.")
                        else:
                            st.session_state.liquidity_rows[idx] = {
                                "nome": el_nome.strip(),
                                "tipo": el_tipo,
                                "importo": float(el_importo),
                                "valuta": (el_valuta or "EUR").strip().upper(),
                            }
                            save_liquidity(st.session_state.liquidity_rows)
                            st.rerun()
                    if l_del_clicked:
                        st.session_state.liquidity_rows.pop(idx)
                        save_liquidity(st.session_state.liquidity_rows)
                        st.rerun()

    st.markdown("---")
    st.markdown(
        '<p style="font-family: JetBrains Mono, monospace; font-size:0.7rem; color:#A09A8E !important;">v2.1 · Streamlit + Plotly + Claude</p>',
        unsafe_allow_html=True,
    )

# ── Landing state ────────────────────────────────────────────────────────────

if not uploaded_files:
    st.markdown(
        """
        <div class="hero-container">
            <div class="hero-logo">
                <span class="mat" style="font-size:2.6rem;
                    font-variation-settings:'FILL' 0,'wght' 400,'GRAD' 0,'opsz' 48;">
                    insights
                </span>
            </div>
            <div class="hero-title">The Portfolio<br><em>Almanac</em></div>
            <div class="hero-subtitle">
                Un'analisi tipografica del tuo capitale —<br>
                allocazione, composizione, esposizione e rischio,<br>
                composti come pagine di un almanacco finanziario.
            </div>
            <div class="hero-hint">
                <span class="mat arrow" style="font-size:0.9rem;">west</span>
                Carica un file per iniziare
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.stop()

# ── Load & process (csv-data-summarizer + xlsx skill) ───────────────────────

def read_file(uploaded) -> pd.DataFrame:
    """
    Robust file reader per CSV e Excel.
    CSV: prova separatori comuni e encodings multipli.
    Excel: legge il primo foglio, mostra un selettore se ci sono più fogli.
    """
    name = uploaded.name.lower()

    if name.endswith(".csv") or name.endswith(".tsv"):
        separators = [",", ";", "\t", "|"]
        encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]
        known = set(HEADER_MAP.keys())
        best: tuple | None = None   # (df, score)
        last_err = None
        for enc in encodings:
            for sep in separators:
                # Try with auto-header and also skipping up to 10 rows
                for skip in [0] + list(range(1, 11)):
                    try:
                        uploaded.seek(0)
                        df = pd.read_csv(
                            uploaded, sep=sep, encoding=enc,
                            engine="python", skiprows=skip,
                        )
                        if df.shape[1] < 2:
                            continue
                        score = sum(
                            1 for c in df.columns
                            if _norm_key(c) in known
                        )
                        if best is None or score > best[1]:
                            best = (df, score)
                        if score >= 2:      # good enough — stop searching
                            return df, None
                    except Exception as e:
                        last_err = e
        if best is not None:
            return best[0], None
        return None, f"Impossibile leggere il CSV: {last_err}"

    else:  # .xlsx / .xls
        try:
            uploaded.seek(0)
            xl = pd.ExcelFile(uploaded)
            sheets = xl.sheet_names
            if len(sheets) > 1:
                sheet = st.sidebar.selectbox(
                    "Foglio Excel",
                    sheets,
                    help="Il file contiene più fogli — scegli quello da analizzare.",
                )
            else:
                sheet = sheets[0]

            # Auto-detect header row: scan first 15 rows for the one that best
            # matches our known column names from HEADER_MAP
            known = set(HEADER_MAP.keys())

            def _col_score(columns) -> int:
                score = 0
                for c in columns:
                    k = _norm_key(c)
                    if k in known:
                        score += 1
                    elif re.sub(r"[€$£]+$", "", k).strip() in known:
                        score += 1
                return score

            best_row = 0
            best_score = 0
            for skip in range(15):
                uploaded.seek(0)
                try:
                    probe = xl.parse(sheet, header=skip, nrows=3)
                    score = _col_score(probe.columns)
                    if score > best_score:
                        best_score = score
                        best_row = skip
                except Exception:
                    break

            uploaded.seek(0)
            df = xl.parse(sheet, header=best_row)
            # Drop fully-empty rows/cols that often appear above real data
            df = df.dropna(how="all").dropna(axis=1, how="all")
            return df, None
        except Exception as e:
            return None, f"Impossibile leggere l'Excel: {e}"


def data_quality_report(raw: pd.DataFrame) -> dict:
    """Genera le statistiche di qualità dati (ispirato a csv-data-summarizer)."""
    total_cells = raw.shape[0] * raw.shape[1]
    missing_cells = raw.isnull().sum().sum()
    missing_pct = (missing_cells / total_cells * 100) if total_cells else 0
    by_col = {
        col: {"count": int(raw[col].isnull().sum()),
              "pct": round(raw[col].isnull().sum() / len(raw) * 100, 1)}
        for col in raw.columns
        if raw[col].isnull().sum() > 0
    }
    return {
        "rows": raw.shape[0],
        "cols": raw.shape[1],
        "missing_cells": int(missing_cells),
        "missing_pct": round(missing_pct, 1),
        "by_col": by_col,
    }


all_raws = []
for uf in uploaded_files:
    raw_i, read_error = read_file(uf)
    if read_error:
        st.error(f"{uf.name}: {read_error}")
        st.stop()
    raw_i["_fonte"] = uf.name
    all_raws.append(raw_i)

raw = pd.concat(all_raws, ignore_index=True) if len(all_raws) > 1 else all_raws[0]

qr = data_quality_report(raw)
raw_copy = raw.copy()
df = process_portfolio(raw_copy)
# _fonte sopravvive a process_portfolio perché non viene toccata
if "_fonte" in df.columns:
    df["Fonte"] = df["_fonte"]
    df = df.drop(columns=["_fonte"])

# ── Merge crypto manuali nel portafoglio ─────────────────────────────────────
crypto_df = crypto_rows_to_df(st.session_state.get("crypto_rows", []))
if not crypto_df.empty:
    df = pd.concat([df, crypto_df], ignore_index=True)

# ── Merge liquidità manuale nel portafoglio ──────────────────────────────────
liquidity_df = liquidity_rows_to_df(st.session_state.get("liquidity_rows", []))
if not liquidity_df.empty:
    df = pd.concat([df, liquidity_df], ignore_index=True)

# Ricalcola i pesi su tutto (file + crypto + liquidità)
if not crypto_df.empty or not liquidity_df.empty:
    total_after = df["Controvalore"].sum()
    df["Peso %"] = (df["Controvalore"] / total_after * 100).round(2) if total_after > 0 else 0

# ── KPI row ──────────────────────────────────────────────────────────────────

total_value = df["Controvalore"].sum()
n_assets = len(df)
has_values = not df["Controvalore"].isna().all()
heaviest = df.loc[df["Controvalore"].idxmax(), "Nome"] if has_values else "N/A"
heaviest_pct = df["Peso %"].max() if has_values else 0

# Truncate long names for display
heaviest_short = (heaviest[:30] + "...") if len(heaviest) > 30 else heaviest

st.markdown(
    '<div class="section-label"><span class="mat" style="font-size:0.9rem; vertical-align:-2px;">dashboard</span>&nbsp; Panoramica</div>',
    unsafe_allow_html=True,
)
k1, k2, k3 = st.columns(3)
with k1:
    st.markdown(
        f"""<div class="glass-card kpi-card kpi-bond">
            <div class="kpi-icon kpi-icon-value" style="background:rgba(31,95,91,0.07);">
                <span class="mat" style="color:#1F5F5B;font-variation-settings:'FILL' 1,'wght' 300,'GRAD' 0,'opsz' 24;">account_balance_wallet</span>
            </div>
            <div class="kpi-label" style="color:#1F5F5B !important;">Controvalore Totale</div>
            <div class="kpi-value" style="color:#1F5F5B !important;">€ {total_value:,.2f}</div>
        </div>""",
        unsafe_allow_html=True,
    )
with k2:
    st.markdown(
        f"""<div class="glass-card kpi-card kpi-etf">
            <div class="kpi-icon kpi-icon-count" style="background:rgba(59,90,54,0.07);">
                <span class="mat" style="color:#3B5A36;font-variation-settings:'FILL' 1,'wght' 300,'GRAD' 0,'opsz' 24;">dataset</span>
            </div>
            <div class="kpi-label" style="color:#3B5A36 !important;">Numero Titoli</div>
            <div class="kpi-value" style="color:#3B5A36 !important;">{n_assets}</div>
        </div>""",
        unsafe_allow_html=True,
    )
with k3:
    st.markdown(
        f"""<div class="glass-card kpi-card kpi-cert">
            <div class="kpi-icon kpi-icon-top" style="background:rgba(107,46,95,0.07);">
                <span class="mat" style="color:#6B2E5F;font-variation-settings:'FILL' 1,'wght' 300,'GRAD' 0,'opsz' 24;">trophy</span>
            </div>
            <div class="kpi-label" style="color:#6B2E5F !important;">Asset Più Pesante</div>
            <div class="kpi-value" style="font-size:0.95rem; line-height:1.4; color:#6B2E5F !important;">{heaviest_short}</div>
            <div class="kpi-sub" style="color:#6B2E5F;">{heaviest_pct:.1f}%</div>
        </div>""",
        unsafe_allow_html=True,
    )

# ── Chromatic allocation bar ─────────────────────────────────────────────────

CLASS_COLORS = {
    "BOND":        ("#1F5F5B", "rgba(31,95,91,0.16)"),
    "ETF":         ("#3B5A36", "rgba(59,90,54,0.16)"),
    "CERTIFICATE": ("#6B2E5F", "rgba(107,46,95,0.16)"),
    "COMMODITY":   ("#B45309", "rgba(180,83,9,0.16)"),
    "AZIONE":      ("#1E3A8A", "rgba(30,58,138,0.16)"),
    "CRYPTO":      ("#C9893A", "rgba(201,137,58,0.16)"),
    "LIQUIDITÀ":   ("#7C7C5A", "rgba(124,124,90,0.16)"),
    "ALTRO":       ("#6B6258", "rgba(107,98,88,0.16)"),
}

if has_values:
    class_totals = df.groupby("Classe")["Controvalore"].sum()
    total_ctv = class_totals.sum()
    segments = []
    for cls, val in class_totals.items():
        pct = val / total_ctv * 100
        col_l, col_f = CLASS_COLORS.get(cls, ("#A09A8E", "rgba(74,85,104,0.18)"))
        segments.append((cls, pct, col_l, col_f))
    segments.sort(key=lambda x: -x[1])

    bar_html = '<div style="display:flex; width:100%; height:6px; border-radius:3px; overflow:hidden; margin:1rem 0 0.4rem;">'
    for cls, pct, col_l, _ in segments:
        bar_html += f'<div style="width:{pct:.2f}%; background:{col_l}; opacity:0.7;"></div>'
    bar_html += '</div>'

    labels_html = '<div style="display:flex; gap:1.2rem; flex-wrap:wrap; margin-bottom:1.5rem;">'
    for cls, pct, col_l, _ in segments:
        labels_html += (
            f'<span style="display:inline-flex; align-items:center; gap:0.35rem; '
            f'font-family:JetBrains Mono,monospace; font-size:0.7rem; color:{col_l};">'
            f'<span style="display:inline-block; width:8px; height:8px; border-radius:1px; '
            f'background:{col_l};"></span>{cls}&nbsp;<span style="opacity:0.5;">{pct:.1f}%</span></span>'
        )
    labels_html += '</div>'

    st.markdown(bar_html + labels_html, unsafe_allow_html=True)

# ── Risk metrics & insights ─────────────────────────────────────────────────

risk_metrics = compute_risk_metrics(df)
portfolio_insights = generate_insights(df, risk_metrics) if risk_metrics else []

if portfolio_insights:
    insight_html = "".join(f"<li>{i}</li>" for i in portfolio_insights)
    st.markdown(
        f"""<div class="glass-card" style="padding:1rem 1.4rem; margin-bottom:1rem;">
            <div style="display:flex; align-items:center; gap:0.4rem; margin-bottom:0.6rem;">
                <span class="mat" style="font-size:1.1rem; color:#6B2E5F;
                    font-variation-settings:'FILL' 1,'wght' 300,'GRAD' 0,'opsz' 24;">psychology</span>
                <span style="font-size:0.78rem; font-weight:600; letter-spacing:0.05em;
                    text-transform:uppercase; color:#6B2E5F;">Insight</span>
            </div>
            <ul style="margin:0; padding-left:1.2rem; font-size:0.88rem; color:#1A1815;
                line-height:1.7; list-style-type:'›  ';">{insight_html}</ul>
        </div>""",
        unsafe_allow_html=True,
    )

st.markdown("<div style='height:1.8rem'></div>", unsafe_allow_html=True)

# ── Claude prompt context (usato dalla tab Analisi e dalla Chat) ────────────

def _portfolio_context(df: pd.DataFrame, risk: dict | None, goals: str | None = None) -> str:
    """Costruisce un contesto testuale compatto del portafoglio per Claude."""
    total = df["Controvalore"].sum()
    lines = [f"# Portafoglio — Controvalore totale: € {total:,.2f}", f"Numero titoli: {len(df)}", ""]

    if goals and goals.strip():
        lines.append("## Obiettivi dell'utente (file Markdown)")
        lines.append(goals.strip())
        lines.append("")

    by_class = df.groupby("Classe")["Controvalore"].sum().sort_values(ascending=False)
    lines.append("## Allocazione per classe")
    for cls, val in by_class.items():
        pct = val / total * 100 if total else 0
        lines.append(f"- {cls}: € {val:,.2f} ({pct:.2f}%)")
    lines.append("")

    by_sub = df.groupby("Sottotipo")["Controvalore"].sum().sort_values(ascending=False)
    lines.append("## Allocazione per sottotipo")
    for sub, val in by_sub.head(15).items():
        pct = val / total * 100 if total else 0
        lines.append(f"- {sub}: € {val:,.2f} ({pct:.2f}%)")
    lines.append("")

    by_geo = df.groupby("Geografia")["Controvalore"].sum().sort_values(ascending=False)
    lines.append("## Esposizione geografica")
    for geo, val in by_geo.items():
        pct = val / total * 100 if total else 0
        lines.append(f"- {geo}: € {val:,.2f} ({pct:.2f}%)")
    lines.append("")

    lines.append("## Holdings (Nome | Classe | Controvalore | Peso %)")
    cols = [c for c in ("Nome", "Classe", "Controvalore", "Peso %", "Sottotipo", "Geografia") if c in df.columns]
    rows = df[cols].sort_values("Controvalore", ascending=False)
    for _, r in rows.iterrows():
        nome = str(r.get("Nome", ""))[:80]
        ctv = r.get("Controvalore")
        peso = r.get("Peso %")
        ctv_s = f"€ {ctv:,.2f}" if pd.notna(ctv) else "N/D"
        peso_s = f"{peso:.2f}%" if pd.notna(peso) else "N/D"
        lines.append(
            f"- {nome} | {r.get('Classe','')} | {ctv_s} | {peso_s} "
            f"| {r.get('Sottotipo','')} | {r.get('Geografia','')}"
        )

    if risk:
        lines.append("")
        lines.append("## Metriche di rischio")
        for k, v in risk.items():
            if isinstance(v, (int, float)):
                lines.append(f"- {k}: {v:.2f}" if isinstance(v, float) else f"- {k}: {v}")

    return "\n".join(lines)


CHAT_SYSTEM_PROMPT = """Sei un consulente finanziario che assiste l'utente nell'analisi del suo portafoglio.
Rispondi in italiano, in modo conciso e professionale.

Hai accesso al portafoglio completo dell'utente nel messaggio successivo. Se nello stesso messaggio
è presente una sezione "## Obiettivi dell'utente", trattala come la dichiarazione di intenti che guida
ogni tua risposta: orizzonte, target di asset allocation, vincoli, scadenze. Confronta sempre lo stato
attuale con gli obiettivi e segnala scostamenti.

Quando l'utente chiede di:
- ribilanciare per liberare liquidità: proponi vendite specifiche con importi e motivazioni (es. ridurre concentrazione, mantenere diversificazione, vendere prima asset più liquidi/meno strategici)
- valutare il rischio: usa le metriche fornite (HHI, %azionario, %obbligazionario)
- spiegare scelte: rifletti sulla logica dell'allocazione e degli obiettivi (se presenti)

Regole:
- Quando proponi vendite, indica sempre: nome del titolo, importo da vendere in €, peso prima/dopo, motivazione.
- Per liberare liquidità preferisci: 1) ridurre posizioni sovrappesate, 2) vendere asset più liquidi (ETF, azioni, crypto), 3) mantenere diversificazione di classe.
- Non dare mai consigli che assomiglino a "promesse di rendimento". Ricorda che è una simulazione informativa, non consulenza autorizzata.
- Se mancano dati per rispondere, chiedilo invece di inventare.
"""


GOALS_ANALYSIS_PROMPT = """Analizza il portafoglio dell'utente alla luce degli obiettivi che ha dichiarato.

Struttura la risposta in queste sezioni (usa intestazioni markdown `##`):

## Sintesi
2-3 frasi: quanto il portafoglio attuale è coerente con gli obiettivi (allineato / parzialmente / non allineato).

## Scostamenti
Per ciascun obiettivo dichiarato, indica:
- valore target (se specificato dall'utente)
- valore attuale (dal portafoglio)
- delta in € e/o punti %
- valutazione (verde/giallo/rosso)

## Punti di forza
Cosa è già in linea con gli obiettivi (max 3 bullet).

## Aree di intervento
Lista concreta di azioni operative per avvicinare il portafoglio agli obiettivi:
- ogni azione deve avere: cosa fare, importo €, motivazione
- ordinata per priorità (impatto/urgenza)

## Domande aperte
Cosa manca o è ambiguo negli obiettivi per dare una risposta più precisa.

Sii concreto, mai generico. Niente disclaimer lunghi: alla fine una sola riga "_Simulazione informativa, non consulenza._"
"""


# ── Tabs ─────────────────────────────────────────────────────────────────────

tab_charts, tab_table, tab_analysis, tab_chat = st.tabs(
    ["Grafici", "Tabella Dati", "Analisi", "Chat"]
)

with tab_charts:
    c1, c2, c3 = st.columns(3)

    # Donut — asset class
    with c1:
        class_agg = df.groupby("Classe", as_index=False)["Controvalore"].sum()
        fig_class = px.pie(
            class_agg,
            names="Classe",
            values="Controvalore",
            color_discrete_sequence=CHART_PALETTE,
            hole=0.55,
        )
        fig_class.update_traces(
            textinfo="label+percent",
            textposition="auto",
            textfont=dict(size=11, family="Bricolage Grotesque"),
            insidetextorientation="horizontal",
            hovertemplate="<b>%{label}</b><br>€ %{value:,.2f}<br>%{percent}<extra></extra>",
            marker=dict(line=dict(color="#F4ECDF", width=2)),
        )
        fig_class.update_layout(
            **PLOTLY_LAYOUT,
            title_text="Allocazione per Classe",
            annotations=[dict(
                text="Classe",
                x=0.5, y=0.5,
                font=dict(size=13, color="#A09A8E", family="Bricolage Grotesque"),
                showarrow=False,
            )],
        )
        st.plotly_chart(fig_class, use_container_width=True)

    # Donut — sector
    with c2:
        sector_agg = df.groupby("Settore", as_index=False)["Controvalore"].sum()
        fig_sector = px.pie(
            sector_agg,
            names="Settore",
            values="Controvalore",
            color_discrete_sequence=CHART_PALETTE[2:] + CHART_PALETTE[:2],
            hole=0.55,
        )
        fig_sector.update_traces(
            textinfo="label+percent",
            textposition="auto",
            textfont=dict(size=11, family="Bricolage Grotesque"),
            insidetextorientation="horizontal",
            hovertemplate="<b>%{label}</b><br>€ %{value:,.2f}<br>%{percent}<extra></extra>",
            marker=dict(line=dict(color="#F4ECDF", width=2)),
        )
        fig_sector.update_layout(
            **PLOTLY_LAYOUT,
            title_text="Allocazione per Settore",
            annotations=[dict(
                text="Settore",
                x=0.5, y=0.5,
                font=dict(size=13, color="#A09A8E", family="Bricolage Grotesque"),
                showarrow=False,
            )],
        )
        st.plotly_chart(fig_sector, use_container_width=True)

    # Donut — geography
    with c3:
        geo_agg = df.groupby("Geografia", as_index=False)["Controvalore"].sum()
        fig_geo = px.pie(
            geo_agg,
            names="Geografia",
            values="Controvalore",
            color_discrete_sequence=CHART_PALETTE[4:] + CHART_PALETTE[:4],
            hole=0.55,
        )
        fig_geo.update_traces(
            textinfo="label+percent",
            textposition="auto",
            textfont=dict(size=11, family="Bricolage Grotesque"),
            insidetextorientation="horizontal",
            hovertemplate="<b>%{label}</b><br>€ %{value:,.2f}<br>%{percent}<extra></extra>",
            marker=dict(line=dict(color="#F4ECDF", width=2)),
        )
        fig_geo.update_layout(
            **PLOTLY_LAYOUT,
            title_text="Allocazione Geografica",
            annotations=[dict(
                text="Geo",
                x=0.5, y=0.5,
                font=dict(size=13, color="#A09A8E", family="Bricolage Grotesque"),
                showarrow=False,
            )],
        )
        st.plotly_chart(fig_geo, use_container_width=True)

    # ── Analisi diversificazione (financial-analyst) ────────────────────────
    df_bond = df[df["Classe"] == "BOND"].dropna(subset=["Controvalore"])
    df_equity = df[df["Classe"].isin(["ETF", "AZIONE"])].dropna(subset=["Controvalore"])
    df_other = df[~df["Classe"].isin(["BOND", "ETF", "AZIONE"])].dropna(subset=["Controvalore"])

    has_bond = len(df_bond) > 0
    has_equity = len(df_equity) > 0

    if has_bond or has_equity:
        st.markdown(
            '<div class="section-label" style="margin-top:1.5rem;">'
            '<span class="mat" style="font-size:0.9rem; vertical-align:-2px;">analytics</span>'
            '&nbsp; Analisi diversificazione</div>',
            unsafe_allow_html=True,
        )

        # ── KPI riepilogo macro-allocazione ──
        bond_total_val = df_bond["Controvalore"].sum() if has_bond else 0
        eq_total_val = df_equity["Controvalore"].sum() if has_equity else 0
        other_total_val = df_other["Controvalore"].sum() if len(df_other) > 0 else 0
        ptf_total = bond_total_val + eq_total_val + other_total_val
        bond_pct = (bond_total_val / ptf_total * 100) if ptf_total > 0 else 0
        eq_pct = (eq_total_val / ptf_total * 100) if ptf_total > 0 else 0
        other_pct = (other_total_val / ptf_total * 100) if ptf_total > 0 else 0

        alloc_cols = st.columns(3 if other_total_val > 0 else 2)
        with alloc_cols[0]:
            st.markdown(
                f"""<div class="glass-card" style="text-align:center; padding:1rem 1.2rem;">
                    <span class="mat" style="font-size:1.3rem; color:#1F5F5B;
                        font-variation-settings:'FILL' 1,'wght' 300,'GRAD' 0,'opsz' 24;">savings</span>
                    <div style="font-size:0.72rem; font-weight:600; letter-spacing:0.06em;
                        text-transform:uppercase; color:#A09A8E !important; margin:0.3rem 0 0.2rem;">Obbligazionario</div>
                    <div style="font-family:'JetBrains Mono',monospace; font-size:1.8rem; font-weight:700;
                        color:#1F5F5B !important;">{bond_pct:.1f}%</div>
                    <div style="font-family:'JetBrains Mono',monospace; font-size:0.8rem;
                        color:#6B6258 !important;">€ {bond_total_val:,.0f}</div>
                </div>""",
                unsafe_allow_html=True,
            )
        with alloc_cols[1]:
            st.markdown(
                f"""<div class="glass-card" style="text-align:center; padding:1rem 1.2rem;">
                    <span class="mat" style="font-size:1.3rem; color:#3B5A36;
                        font-variation-settings:'FILL' 1,'wght' 300,'GRAD' 0,'opsz' 24;">trending_up</span>
                    <div style="font-size:0.72rem; font-weight:600; letter-spacing:0.06em;
                        text-transform:uppercase; color:#A09A8E !important; margin:0.3rem 0 0.2rem;">Azionario</div>
                    <div style="font-family:'JetBrains Mono',monospace; font-size:1.8rem; font-weight:700;
                        color:#3B5A36 !important;">{eq_pct:.1f}%</div>
                    <div style="font-family:'JetBrains Mono',monospace; font-size:0.8rem;
                        color:#6B6258 !important;">€ {eq_total_val:,.0f}</div>
                </div>""",
                unsafe_allow_html=True,
            )
        if other_total_val > 0:
            with alloc_cols[2]:
                st.markdown(
                    f"""<div class="glass-card" style="text-align:center; padding:1rem 1.2rem;">
                        <span class="mat" style="font-size:1.3rem; color:#B45309;
                            font-variation-settings:'FILL' 1,'wght' 300,'GRAD' 0,'opsz' 24;">category</span>
                        <div style="font-size:0.72rem; font-weight:600; letter-spacing:0.06em;
                            text-transform:uppercase; color:#A09A8E !important; margin:0.3rem 0 0.2rem;">Altro</div>
                        <div style="font-family:'JetBrains Mono',monospace; font-size:1.8rem; font-weight:700;
                            color:#B45309 !important;">{other_pct:.1f}%</div>
                        <div style="font-family:'JetBrains Mono',monospace; font-size:0.8rem;
                            color:#6B6258 !important;">€ {other_total_val:,.0f}</div>
                    </div>""",
                    unsafe_allow_html=True,
                )

        st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)

        # ── Row 1: Sottotipo donut (per-sottocategoria) ──
        d1, d2 = st.columns(2)

        if has_bond:
            with d1:
                bond_total = df_bond["Controvalore"].sum()
                bond_sub = df_bond.groupby("Sottotipo", as_index=False)["Controvalore"].sum()
                bond_sub = bond_sub.sort_values("Controvalore", ascending=False)
                fig_bond_sub = px.pie(
                    bond_sub,
                    names="Sottotipo",
                    values="Controvalore",
                    color_discrete_sequence=CHART_PALETTE,
                    hole=0.55,
                )
                fig_bond_sub.update_traces(
                    textinfo="label+percent",
                    textposition="outside",
                    textfont=dict(size=10, family="Bricolage Grotesque"),
                    insidetextorientation="horizontal",
                    hovertemplate="<b>%{label}</b><br>€ %{value:,.2f}<br>%{percent}<extra></extra>",
                    marker=dict(line=dict(color="#F4ECDF", width=2)),
                )
                bond_sub_layout = {k: v for k, v in PLOTLY_LAYOUT.items() if k != "legend"}
                fig_bond_sub.update_layout(
                    **bond_sub_layout,
                    title_text=f"Obbligazionario — € {bond_total:,.0f}",
                    showlegend=True,
                    legend=dict(font=dict(size=10, color="#6B6258"), bgcolor="rgba(0,0,0,0)"),
                    annotations=[dict(
                        text="Bond", x=0.5, y=0.5,
                        font=dict(size=13, color="#A09A8E", family="Bricolage Grotesque"),
                        showarrow=False,
                    )],
                )
                st.plotly_chart(fig_bond_sub, use_container_width=True)

        if has_equity:
            with d2:
                eq_total = df_equity["Controvalore"].sum()
                eq_sub = df_equity.groupby("Sottotipo", as_index=False)["Controvalore"].sum()
                eq_sub = eq_sub.sort_values("Controvalore", ascending=False)
                fig_eq_sub = px.pie(
                    eq_sub,
                    names="Sottotipo",
                    values="Controvalore",
                    color_discrete_sequence=CHART_PALETTE[2:] + CHART_PALETTE[:2],
                    hole=0.55,
                )
                fig_eq_sub.update_traces(
                    textinfo="label+percent",
                    textposition="outside",
                    textfont=dict(size=10, family="Bricolage Grotesque"),
                    insidetextorientation="horizontal",
                    hovertemplate="<b>%{label}</b><br>€ %{value:,.2f}<br>%{percent}<extra></extra>",
                    marker=dict(line=dict(color="#F4ECDF", width=2)),
                )
                eq_sub_layout = {k: v for k, v in PLOTLY_LAYOUT.items() if k != "legend"}
                fig_eq_sub.update_layout(
                    **eq_sub_layout,
                    title_text=f"Azionario — € {eq_total:,.0f}",
                    showlegend=True,
                    legend=dict(font=dict(size=10, color="#6B6258"), bgcolor="rgba(0,0,0,0)"),
                    annotations=[dict(
                        text="Equity", x=0.5, y=0.5,
                        font=dict(size=13, color="#A09A8E", family="Bricolage Grotesque"),
                        showarrow=False,
                    )],
                )
                st.plotly_chart(fig_eq_sub, use_container_width=True)

        # ── Row 2: Dettaglio singoli titoli per classe ──
        st.markdown(
            '<div class="section-label" style="margin-top:1rem;">'
            '<span class="mat" style="font-size:0.9rem; vertical-align:-2px;">list_alt</span>'
            '&nbsp; Dettaglio titoli per classe</div>',
            unsafe_allow_html=True,
        )
        e1, e2 = st.columns(2)

        if has_bond:
            with e1:
                df_b = df_bond.sort_values("Controvalore", ascending=True).copy()
                fig_b_bar = px.bar(
                    df_b,
                    x="Controvalore",
                    y="Nome",
                    orientation="h",
                    color="Sottotipo",
                    color_discrete_sequence=CHART_PALETTE,
                    custom_data=["ISIN", "Peso %", "Sottotipo"],
                )
                fig_b_bar.update_traces(
                    hovertemplate=(
                        "<b>%{y}</b><br>ISIN: %{customdata[0]}<br>"
                        "€ %{x:,.2f} · %{customdata[1]:.1f}%<br>"
                        "%{customdata[2]}<extra></extra>"
                    ),
                    marker=dict(cornerradius=4),
                )
                b_layout = {k: v for k, v in PLOTLY_LAYOUT.items() if k != "legend"}
                fig_b_bar.update_layout(
                    **b_layout,
                    title_text="Titoli Obbligazionari",
                    height=max(300, len(df_b) * 32),
                    xaxis=dict(gridcolor="rgba(26,24,21,0.06)", zeroline=False,
                               tickfont=dict(family="JetBrains Mono", size=9, color="#A09A8E")),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)",
                               tickfont=dict(family="Bricolage Grotesque", size=10, color="#6B6258")),
                    showlegend=True,
                    legend=dict(orientation="h", y=-0.15, x=0.5, xanchor="center",
                                font=dict(size=10, color="#6B6258"), bgcolor="rgba(0,0,0,0)"),
                )
                st.plotly_chart(fig_b_bar, use_container_width=True)

        if has_equity:
            with e2:
                df_e = df_equity.sort_values("Controvalore", ascending=True).copy()
                fig_e_bar = px.bar(
                    df_e,
                    x="Controvalore",
                    y="Nome",
                    orientation="h",
                    color="Sottotipo",
                    color_discrete_sequence=CHART_PALETTE[2:] + CHART_PALETTE[:2],
                    custom_data=["ISIN", "Peso %", "Sottotipo"],
                )
                fig_e_bar.update_traces(
                    hovertemplate=(
                        "<b>%{y}</b><br>ISIN: %{customdata[0]}<br>"
                        "€ %{x:,.2f} · %{customdata[1]:.1f}%<br>"
                        "%{customdata[2]}<extra></extra>"
                    ),
                    marker=dict(cornerradius=4),
                )
                e_layout = {k: v for k, v in PLOTLY_LAYOUT.items() if k != "legend"}
                fig_e_bar.update_layout(
                    **e_layout,
                    title_text="Titoli Azionari",
                    height=max(300, len(df_e) * 32),
                    xaxis=dict(gridcolor="rgba(26,24,21,0.06)", zeroline=False,
                               tickfont=dict(family="JetBrains Mono", size=9, color="#A09A8E")),
                    yaxis=dict(gridcolor="rgba(0,0,0,0)",
                               tickfont=dict(family="Bricolage Grotesque", size=10, color="#6B6258")),
                    showlegend=True,
                    legend=dict(orientation="h", y=-0.15, x=0.5, xanchor="center",
                                font=dict(size=10, color="#6B6258"), bgcolor="rgba(0,0,0,0)"),
                )
                st.plotly_chart(fig_e_bar, use_container_width=True)

    # Horizontal bar — top holdings
    df_sorted = df.dropna(subset=["Controvalore"]).sort_values("Controvalore", ascending=True)

    # Color map per class for the bar chart
    class_color = {cls: CHART_PALETTE[i % len(CHART_PALETTE)] for i, cls in enumerate(df["Classe"].unique())}

    fig_bar = go.Figure()
    for cls in df_sorted["Classe"].unique():
        subset = df_sorted[df_sorted["Classe"] == cls]
        fig_bar.add_trace(go.Bar(
            x=subset["Controvalore"],
            y=subset["Nome"],
            orientation="h",
            name=cls,
            marker=dict(
                color=class_color.get(cls, CHART_PALETTE[0]),
                line=dict(width=0),
                cornerradius=4,
            ),
            customdata=list(zip(subset["ISIN"], subset["Peso %"], subset["Classe"])),
            hovertemplate=(
                "<b>%{y}</b><br>"
                "ISIN: %{customdata[0]}<br>"
                "Controvalore: € %{x:,.2f}<br>"
                "Peso: %{customdata[1]:.1f}%<br>"
                "Classe: %{customdata[2]}"
                "<extra></extra>"
            ),
        ))

    bar_height = max(400, len(df_sorted) * 36)
    bar_layout = {k: v for k, v in PLOTLY_LAYOUT.items() if k != "legend"}
    fig_bar.update_layout(
        **bar_layout,
        title_text="Holdings per Controvalore",
        height=bar_height,
        barmode="stack",
        xaxis=dict(
            gridcolor="rgba(26,24,21,0.06)",
            zeroline=False,
            tickfont=dict(family="JetBrains Mono", size=10, color="#A09A8E"),
        ),
        yaxis=dict(
            gridcolor="rgba(0,0,0,0)",
            tickfont=dict(family="Bricolage Grotesque", size=11, color="#6B6258"),
        ),
        showlegend=True,
        legend=dict(orientation="h", y=-0.12, x=0.5, xanchor="center",
                    font=dict(size=11, color="#6B6258"), bgcolor="rgba(0,0,0,0)"),
    )
    st.plotly_chart(fig_bar, use_container_width=True)

with tab_table:
    display_cols = ["Nome", "ISIN", "Classe", "Settore", "Geografia", "Quantità", "Prezzo Acquisto", "Controvalore", "Peso %"]
    display_cols = [c for c in display_cols if c in df.columns]
    df_display = df[display_cols].copy()

    def highlight_missing(val):
        if pd.isna(val) or str(val).strip() in ("", "N/A", "n/a", "n/d"):
            return "background-color: rgba(180,83,9,0.10); color: #B45309;"
        return ""

    styled = df_display.style.applymap(highlight_missing).format(
        {
            "Controvalore": "€ {:,.2f}",
            "Prezzo Acquisto": "€ {:,.2f}",
            "Peso %": "{:.2f}%",
        },
        na_rep="— N/D",
    )
    st.dataframe(styled, use_container_width=True, height=480)

    st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)
    col_dl1, col_dl2, _ = st.columns([1, 1, 3])
    with col_dl1:
        csv_bytes = df_display.to_csv(index=False).encode("utf-8")
        st.download_button("download  CSV", csv_bytes, "portafoglio_analizzato.csv", "text/csv")
    with col_dl2:
        excel_summary = {
            "total_value": total_value,
            "n_assets": n_assets,
            "heaviest": heaviest,
            "heaviest_pct": heaviest_pct,
            "bond_pct": risk_metrics.get("bond_pct", 0),
            "eq_pct": risk_metrics.get("eq_pct", 0),
            "other_pct": risk_metrics.get("other_pct", 0),
            **{k: v for k, v in risk_metrics.items() if k in (
                "hhi", "top5_pct", "n_subtypes", "diversification_rating"
            )},
        }
        xlsx_bytes = to_excel_styled(df_display, summary=excel_summary)
        st.download_button(
            "download  Excel",
            xlsx_bytes,
            "portafoglio_analizzato.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

with tab_analysis:
    if not risk_metrics:
        st.info("Dati insufficienti per l'analisi di rischio.")
    else:
        # ── Risk KPI cards ──
        st.markdown(
            '<div class="section-label"><span class="mat" style="font-size:0.9rem; vertical-align:-2px;">'
            'security</span>&nbsp; Metriche di Rischio</div>',
            unsafe_allow_html=True,
        )
        r1, r2, r3, r4 = st.columns(4)
        with r1:
            st.markdown(
                f"""<div class="glass-card kpi-card">
                    <div class="kpi-label">Indice HHI</div>
                    <div class="kpi-value" style="font-size:1.6rem;">{risk_metrics['hhi']:.4f}</div>
                    <div class="kpi-sub" style="color:{risk_metrics['diversification_color']};">
                        {risk_metrics['diversification_rating']}</div>
                </div>""",
                unsafe_allow_html=True,
            )
        with r2:
            st.markdown(
                f"""<div class="glass-card kpi-card">
                    <div class="kpi-label">Top-5 Concentrazione</div>
                    <div class="kpi-value" style="font-size:1.6rem;">{risk_metrics['top5_pct']:.1f}%</div>
                    <div class="kpi-sub" style="color:{'#C9893A' if risk_metrics['top5_pct'] > 80 else '#B45309' if risk_metrics['top5_pct'] > 60 else '#3B5A36'};">
                        {'Alto' if risk_metrics['top5_pct'] > 80 else 'Moderato' if risk_metrics['top5_pct'] > 60 else 'Basso'}</div>
                </div>""",
                unsafe_allow_html=True,
            )
        with r3:
            st.markdown(
                f"""<div class="glass-card kpi-card">
                    <div class="kpi-label">Rischio Single-Name</div>
                    <div class="kpi-value" style="font-size:1.6rem;">{risk_metrics['max_weight']:.1f}%</div>
                    <div class="kpi-sub" style="color:{risk_metrics['concentration_color']};">
                        {risk_metrics['concentration_rating']}</div>
                </div>""",
                unsafe_allow_html=True,
            )
        with r4:
            st.markdown(
                f"""<div class="glass-card kpi-card">
                    <div class="kpi-label">Sottocategorie Attive</div>
                    <div class="kpi-value" style="font-size:1.6rem;">{risk_metrics['n_subtypes']}</div>
                    <div class="kpi-sub" style="color:#6B6258;">peso &gt; 2%</div>
                </div>""",
                unsafe_allow_html=True,
            )

        st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

        # ── Analisi vs Obiettivi (file .md caricato dalla sidebar) ──
        st.markdown(
            '<div class="section-label">'
            '<span class="mat" style="font-size:0.9rem; vertical-align:-2px;">flag</span>'
            '&nbsp; Analisi vs Obiettivi</div>',
            unsafe_allow_html=True,
        )

        goals_text = st.session_state.get("goals_text", "")
        if not goals_text:
            st.markdown(
                '<p style="font-size:0.86rem; color:#6B6258; line-height:1.6;">'
                "Carica un file <code>.md</code> dalla sidebar (sezione "
                "<strong>Obiettivi</strong>) per descrivere orizzonte temporale, "
                "asset allocation target, vincoli di liquidità e regole di "
                "ribilanciamento. Verranno usati per generare un'analisi mirata "
                "del portafoglio.</p>",
                unsafe_allow_html=True,
            )
        else:
            with st.expander("Obiettivi correnti", expanded=False):
                st.markdown(goals_text)

            _ga_api_key = os.environ.get("ANTHROPIC_API_KEY")
            if not _ga_api_key:
                try:
                    _ga_api_key = st.secrets["ANTHROPIC_API_KEY"]
                except Exception:
                    _ga_api_key = None

            ga_cols = st.columns([1, 1, 3])
            with ga_cols[0]:
                ga_run = st.button(
                    "Genera analisi", key="btn_goals_analysis",
                    use_container_width=True,
                    disabled=_ga_api_key is None,
                )
            with ga_cols[1]:
                if st.session_state.get("goals_analysis"):
                    if st.button(
                        "Pulisci analisi", key="btn_clear_goals_analysis",
                        use_container_width=True,
                    ):
                        st.session_state.goals_analysis = ""
                        st.rerun()

            if _ga_api_key is None:
                st.markdown(
                    '<p style="font-size:0.78rem; color:#B45309; margin-top:0.5rem;">'
                    "API key Anthropic mancante: imposta <code>ANTHROPIC_API_KEY</code> "
                    "per generare l'analisi.</p>",
                    unsafe_allow_html=True,
                )

            if ga_run and _ga_api_key:
                try:
                    from anthropic import Anthropic
                    _ga_client = Anthropic(api_key=_ga_api_key)
                    _ga_ctx = _portfolio_context(df, risk_metrics, goals_text)
                    _ga_placeholder = st.empty()
                    _ga_full = ""
                    with _ga_client.messages.stream(
                        model="claude-sonnet-4-6",
                        max_tokens=2000,
                        system=GOALS_ANALYSIS_PROMPT,
                        messages=[{
                            "role": "user",
                            "content": [{
                                "type": "text",
                                "text": (
                                    "Ecco lo stato del portafoglio e gli obiettivi:\n\n"
                                    f"{_ga_ctx}"
                                ),
                                "cache_control": {"type": "ephemeral"},
                            }],
                        }],
                    ) as _ga_stream:
                        for _chunk in _ga_stream.text_stream:
                            _ga_full += _chunk
                            _ga_placeholder.markdown(_ga_full + "▌")
                    _ga_placeholder.markdown(_ga_full)
                    st.session_state.goals_analysis = _ga_full
                except ImportError:
                    st.error("Pacchetto `anthropic` non installato. Esegui: `pip3 install anthropic`")
                except Exception as e:
                    st.error(f"Errore chiamata Claude: {e}")
            elif st.session_state.get("goals_analysis"):
                st.markdown(st.session_state.goals_analysis)

        st.markdown("<div style='height:1.5rem'></div>", unsafe_allow_html=True)

        # ── Weight distribution chart ──
        st.markdown(
            '<div class="section-label"><span class="mat" style="font-size:0.9rem; vertical-align:-2px;">'
            'bar_chart</span>&nbsp; Distribuzione Pesi per Sottocategoria</div>',
            unsafe_allow_html=True,
        )
        subtype_agg = df.groupby(["Sottotipo", "Classe"], as_index=False)["Controvalore"].sum()
        subtype_agg["Peso %"] = (subtype_agg["Controvalore"] / total_value * 100).round(2)
        subtype_agg = subtype_agg.sort_values("Peso %", ascending=True)

        class_color_map = {cls: CHART_PALETTE[i % len(CHART_PALETTE)] for i, cls in enumerate(df["Classe"].unique())}
        fig_sub = go.Figure()
        for cls in subtype_agg["Classe"].unique():
            sub = subtype_agg[subtype_agg["Classe"] == cls]
            fig_sub.add_trace(go.Bar(
                x=sub["Peso %"], y=sub["Sottotipo"], orientation="h", name=cls,
                marker=dict(color=class_color_map.get(cls, CHART_PALETTE[0]), cornerradius=4),
                hovertemplate="<b>%{y}</b><br>%{x:.1f}%<br>€ %{customdata:,.0f}<extra></extra>",
                customdata=sub["Controvalore"],
            ))
        sub_layout = {k: v for k, v in PLOTLY_LAYOUT.items() if k != "legend"}
        fig_sub.update_layout(
            **sub_layout,
            height=max(350, len(subtype_agg) * 32),
            barmode="stack",
            xaxis=dict(title="Peso %", gridcolor="rgba(26,24,21,0.06)", zeroline=False,
                       tickfont=dict(family="JetBrains Mono", size=10, color="#A09A8E")),
            yaxis=dict(gridcolor="rgba(0,0,0,0)",
                       tickfont=dict(family="Bricolage Grotesque", size=10, color="#6B6258")),
            showlegend=True,
            legend=dict(orientation="h", y=-0.15, x=0.5, xanchor="center",
                        font=dict(size=10, color="#6B6258"), bgcolor="rgba(0,0,0,0)"),
        )
        st.plotly_chart(fig_sub, use_container_width=True)

        # ── Data quality: correlation & distribution (csv-data-summarizer) ──
        st.markdown(
            '<div class="section-label" style="margin-top:1.5rem;">'
            '<span class="mat" style="font-size:0.9rem; vertical-align:-2px;">query_stats</span>'
            '&nbsp; Analisi Distribuzione</div>',
            unsafe_allow_html=True,
        )
        dq1, dq2 = st.columns(2)

        with dq1:
            # Holding size distribution histogram
            df_valid = df.dropna(subset=["Controvalore"])
            fig_hist = px.histogram(
                df_valid, x="Controvalore", nbins=min(15, max(5, len(df_valid) // 2)),
                color_discrete_sequence=[CHART_PALETTE[0]],
            )
            fig_hist.update_layout(
                **PLOTLY_LAYOUT,
                title_text="Distribuzione Controvalore",
                xaxis=dict(title="Controvalore (€)", gridcolor="rgba(26,24,21,0.06)",
                           tickfont=dict(family="JetBrains Mono", size=10, color="#A09A8E")),
                yaxis=dict(title="Frequenza", gridcolor="rgba(26,24,21,0.06)",
                           tickfont=dict(family="JetBrains Mono", size=10, color="#A09A8E")),
                height=350,
            )
            fig_hist.update_traces(marker_line_width=0)
            st.plotly_chart(fig_hist, use_container_width=True)

        with dq2:
            # Categorical breakdown
            class_counts = df["Classe"].value_counts().reset_index()
            class_counts.columns = ["Classe", "Conteggio"]
            fig_cat = px.bar(
                class_counts, x="Classe", y="Conteggio",
                color="Classe", color_discrete_sequence=CHART_PALETTE,
            )
            fig_cat.update_layout(
                **PLOTLY_LAYOUT,
                title_text="Conteggio per Classe",
                xaxis=dict(gridcolor="rgba(0,0,0,0)",
                           tickfont=dict(family="Bricolage Grotesque", size=10, color="#6B6258")),
                yaxis=dict(gridcolor="rgba(26,24,21,0.06)",
                           tickfont=dict(family="JetBrains Mono", size=10, color="#A09A8E")),
                showlegend=False,
                height=350,
            )
            fig_cat.update_traces(marker=dict(cornerradius=4))
            st.plotly_chart(fig_cat, use_container_width=True)

        # ── Outlier detection ──
        if len(df_valid) > 3:
            mean_ctv = df_valid["Controvalore"].mean()
            std_ctv = df_valid["Controvalore"].std()
            outliers = df_valid[df_valid["Controvalore"] > mean_ctv + 2 * std_ctv]
            if len(outliers) > 0:
                st.markdown(
                    '<div class="section-label" style="margin-top:1rem;">'
                    '<span class="mat" style="font-size:0.9rem; vertical-align:-2px;">warning</span>'
                    '&nbsp; Posizioni Anomale (&gt;2σ dalla media)</div>',
                    unsafe_allow_html=True,
                )
                outlier_display = outliers[["Nome", "Classe", "Controvalore", "Peso %"]].copy()
                st.dataframe(
                    outlier_display.style.format({"Controvalore": "€ {:,.2f}", "Peso %": "{:.2f}%"}),
                    use_container_width=True, hide_index=True,
                )

        # ── What-If Scenario Simulator ──
        st.markdown(
            '<div class="section-label" style="margin-top:2rem;">'
            '<span class="mat" style="font-size:0.9rem; vertical-align:-2px;">science</span>'
            '&nbsp; Simulatore What-If</div>',
            unsafe_allow_html=True,
        )
        with st.expander("Apri simulatore scenari", expanded=False):
            scenario_type = st.radio(
                "Tipo di scenario",
                ["Vendita titolo", "Aggiunta importo", "Stress test"],
                horizontal=True,
            )

            if scenario_type == "Vendita titolo":
                sell_name = st.selectbox(
                    "Titolo da vendere",
                    df["Nome"].dropna().tolist(),
                    key="sell_name",
                )
                if st.button("Simula vendita", key="btn_sell"):
                    df_sim = df[df["Nome"] != sell_name].copy()
                    sim_total = df_sim["Controvalore"].sum()
                    df_sim["Peso %"] = (df_sim["Controvalore"] / sim_total * 100).round(2) if sim_total > 0 else 0
                    sim_metrics = compute_risk_metrics(df_sim)
                    sc1, sc2, sc3 = st.columns(3)
                    with sc1:
                        delta_val = sim_total - total_value
                        st.metric("Controvalore", f"€ {sim_total:,.2f}", f"€ {delta_val:,.2f}")
                    with sc2:
                        delta_hhi = sim_metrics.get("hhi", 0) - risk_metrics.get("hhi", 0)
                        st.metric("HHI", f"{sim_metrics.get('hhi', 0):.4f}", f"{delta_hhi:+.4f}")
                    with sc3:
                        st.metric("Titoli", f"{len(df_sim)}", f"{len(df_sim) - n_assets}")

            elif scenario_type == "Aggiunta importo":
                add_class = st.selectbox(
                    "Classe asset",
                    sorted(df["Classe"].unique().tolist()),
                    key="add_class",
                )
                add_amount = st.number_input(
                    "Importo da aggiungere (€)", min_value=0.0, value=10000.0, step=1000.0,
                    key="add_amount",
                )
                if st.button("Simula aggiunta", key="btn_add"):
                    new_total = total_value + add_amount
                    cur_bond   = risk_metrics.get("bond_pct", 0)
                    cur_eq     = risk_metrics.get("eq_pct", 0)
                    cur_crypto = risk_metrics.get("crypto_pct", 0)
                    cur_other  = risk_metrics.get("other_pct", 0)
                    if add_class == "BOND":
                        new_bond   = (cur_bond * total_value / 100 + add_amount) / new_total * 100
                        new_eq     = cur_eq * total_value / new_total
                        new_crypto = cur_crypto * total_value / new_total
                        new_other  = max(0.0, 100 - new_bond - new_eq - new_crypto)
                    elif add_class in ("ETF", "AZIONE"):
                        new_eq     = (cur_eq * total_value / 100 + add_amount) / new_total * 100
                        new_bond   = cur_bond * total_value / new_total
                        new_crypto = cur_crypto * total_value / new_total
                        new_other  = max(0.0, 100 - new_bond - new_eq - new_crypto)
                    elif add_class == "CRYPTO":
                        new_crypto = (cur_crypto * total_value / 100 + add_amount) / new_total * 100
                        new_bond   = cur_bond * total_value / new_total
                        new_eq     = cur_eq * total_value / new_total
                        new_other  = max(0.0, 100 - new_bond - new_eq - new_crypto)
                    else:
                        new_other  = (cur_other * total_value / 100 + add_amount) / new_total * 100
                        new_bond   = cur_bond * total_value / new_total
                        new_eq     = cur_eq * total_value / new_total
                        new_crypto = cur_crypto * total_value / new_total

                    pie_names  = ["Bond", "Equity", "Crypto", "Altro"]
                    pie_colors = [CHART_PALETTE[0], CHART_PALETTE[1], CHART_PALETTE[5], CHART_PALETTE[7]]
                    sc1, sc2 = st.columns(2)
                    with sc1:
                        st.markdown("**Prima**")
                        fig_before = px.pie(
                            names=pie_names,
                            values=[cur_bond, cur_eq, cur_crypto, cur_other],
                            hole=0.5, color_discrete_sequence=pie_colors,
                        )
                        fig_before.update_layout(**PLOTLY_LAYOUT, height=250, showlegend=True,
                                                 legend=dict(font=dict(size=10)))
                        fig_before.update_traces(textinfo="label+percent", textfont=dict(size=10))
                        st.plotly_chart(fig_before, use_container_width=True)
                    with sc2:
                        st.markdown("**Dopo**")
                        fig_after = px.pie(
                            names=pie_names,
                            values=[new_bond, new_eq, new_crypto, new_other],
                            hole=0.5, color_discrete_sequence=pie_colors,
                        )
                        fig_after.update_layout(**PLOTLY_LAYOUT, height=250, showlegend=True,
                                                legend=dict(font=dict(size=10)))
                        fig_after.update_traces(textinfo="label+percent", textfont=dict(size=10))
                        st.plotly_chart(fig_after, use_container_width=True)

            else:  # Stress test
                stress_pct = st.slider(
                    "Variazione azionario (%)", min_value=-50, max_value=50, value=-20,
                    key="stress_pct",
                )
                crypto_beta = st.slider(
                    "Variazione crypto (% — tipicamente 2×–3× equity in bear market)",
                    min_value=-90, max_value=200, value=-40,
                    key="stress_crypto_pct",
                )
                if st.button("Simula stress test", key="btn_stress"):
                    eq_val     = risk_metrics.get("eq_pct", 0) * total_value / 100
                    crypto_val = risk_metrics.get("crypto_pct", 0) * total_value / 100
                    eq_delta     = eq_val * stress_pct / 100
                    crypto_delta = crypto_val * crypto_beta / 100
                    total_delta  = eq_delta + crypto_delta
                    new_total_stress = total_value + total_delta
                    sc1, sc2, sc3, sc4 = st.columns(4)
                    with sc1:
                        st.metric("Controvalore Attuale", f"€ {total_value:,.2f}")
                    with sc2:
                        st.metric("Impatto Azionario", f"€ {eq_delta:,.2f}",
                                  f"{stress_pct:+d}%")
                    with sc3:
                        st.metric("Impatto Crypto", f"€ {crypto_delta:,.2f}",
                                  f"{crypto_beta:+d}%")
                    with sc4:
                        st.metric("Controvalore Post-Stress", f"€ {new_total_stress:,.2f}",
                                  f"€ {total_delta:,.2f}")

# ── Chat tab (Anthropic Claude) ──────────────────────────────────────────────
# Nota: _portfolio_context, CHAT_SYSTEM_PROMPT e GOALS_ANALYSIS_PROMPT sono
# definiti prima del blocco st.tabs() perché vengono usati anche dalla tab Analisi.

with tab_chat:
    _hdr_left, _hdr_right = st.columns([5, 1])
    with _hdr_left:
        st.markdown(
            '<div class="section-label"><span class="mat" style="font-size:0.9rem; vertical-align:-2px;">'
            'forum</span>&nbsp; Chat con Claude</div>',
            unsafe_allow_html=True,
        )
    with _hdr_right:
        _clear_btn = st.button(
            "Pulisci", help="Cancella la cronologia della chat", use_container_width=True
        )

    st.markdown(
        '<p style="font-size:0.83rem; color:var(--ink-secondary); margin:0.1rem 0 1rem;">'
        "Esempi: <em>«Devo prelevare 30.000 € per la casa, come ribilanci?»</em> · "
        "<em>«Sono troppo concentrato su un titolo?»</em> · "
        "<em>«Quanto pesano le crypto nel mio portafoglio?»</em></p>",
        unsafe_allow_html=True,
    )

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        try:
            api_key = st.secrets["ANTHROPIC_API_KEY"]
        except Exception:
            api_key = None

    if "chat_history" not in st.session_state:
        st.session_state.chat_history = load_chat()

    if _clear_btn:
        st.session_state.chat_history = []
        save_chat([])
        st.rerun()

    if not api_key:
        st.markdown(
            """<div class="glass-card" style="padding:1rem 1.4rem; border-left:3px solid #B45309;">
                <div style="font-size:0.78rem; font-weight:700; letter-spacing:0.06em;
                    text-transform:uppercase; color:#B45309; margin-bottom:0.4rem;">
                    API key mancante
                </div>
                <p style="font-size:0.86rem; color:#1A1815; margin:0; line-height:1.6;">
                    Imposta <code>ANTHROPIC_API_KEY</code> come variabile d'ambiente oppure
                    in <code>.streamlit/secrets.toml</code>:<br>
                    <code style="display:block; margin-top:0.4rem; padding:0.5rem; background:rgba(26,24,21,0.06);
                        border-radius:4px;">ANTHROPIC_API_KEY = "sk-ant-..."</code>
                </p>
            </div>""",
            unsafe_allow_html=True,
        )
    else:
        # ── History messages (scrollable) ─────────────────────────────
        if st.session_state.chat_history:
            with st.container(height=480, border=False):
                turn_idx = 0
                for i, msg in enumerate(st.session_state.chat_history):
                    if msg["role"] == "user":
                        turn_idx += 1
                        if i > 0:
                            st.markdown(
                                f'<div class="chat-turn-sep">Domanda {turn_idx}</div>',
                                unsafe_allow_html=True,
                            )
                    with st.chat_message(msg["role"]):
                        st.markdown(msg["content"])
        else:
            st.markdown(
                '<p style="font-size:0.83rem; color:var(--ink-muted); text-align:center;'
                ' padding:2rem 0; font-style:italic;">Nessuna domanda ancora. Inizia qui sotto.</p>',
                unsafe_allow_html=True,
            )

        user_input = st.chat_input("Chiedi qualcosa al tuo portafoglio…")
        if user_input:
            st.session_state.chat_history.append({"role": "user", "content": user_input})
            save_chat(st.session_state.chat_history)
            with st.chat_message("user"):
                st.markdown(user_input)

            with st.chat_message("assistant"):
                placeholder = st.empty()
                try:
                    from anthropic import Anthropic
                    client = Anthropic(api_key=api_key)

                    portfolio_ctx = _portfolio_context(
                        df, risk_metrics, st.session_state.get("goals_text", "")
                    )
                    api_messages = [
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": (
                                        "Ecco lo stato attuale del mio portafoglio (dati strutturati):\n\n"
                                        f"{portfolio_ctx}"
                                    ),
                                    "cache_control": {"type": "ephemeral"},
                                }
                            ],
                        },
                        {
                            "role": "assistant",
                            "content": "Ricevuto. Sono pronto a rispondere su questo portafoglio.",
                        },
                    ]
                    for m in st.session_state.chat_history:
                        api_messages.append({"role": m["role"], "content": m["content"]})

                    full_text = ""
                    with client.messages.stream(
                        model="claude-sonnet-4-6",
                        max_tokens=1500,
                        system=CHAT_SYSTEM_PROMPT,
                        messages=api_messages,
                    ) as stream:
                        for chunk in stream.text_stream:
                            full_text += chunk
                            placeholder.markdown(full_text + "▌")
                    placeholder.markdown(full_text)
                    st.session_state.chat_history.append({"role": "assistant", "content": full_text})
                    save_chat(st.session_state.chat_history)
                except ImportError:
                    placeholder.error(
                        "Pacchetto `anthropic` non installato. Esegui: `pip3 install anthropic`"
                    )
                except Exception as e:
                    placeholder.error(f"Errore chiamata Claude: {e}")

# ── Financial-Analyst Report ─────────────────────────────────────────────────

if risk_metrics:
    fa = generate_fa_report(df, risk_metrics)

    st.markdown("<div style='height:2rem'></div>", unsafe_allow_html=True)
    st.markdown(
        '<div class="section-label"><span class="mat" style="font-size:0.9rem; vertical-align:-2px;">'
        'person_search</span>&nbsp; Report Financial-Analyst</div>',
        unsafe_allow_html=True,
    )

    # ── Profilo allocazione ───────────────────────────────────────────────────
    prof_name, prof_color, prof_text = fa["alloc_profile"]
    st.markdown(
        f"""<div class="glass-card" style="padding:1.2rem 1.6rem; margin-bottom:0.8rem;
            border-left:3px solid {prof_color}; background:#FFFFFF;">
            <div style="display:flex; align-items:center; gap:0.6rem; margin-bottom:0.5rem;">
                <span style="font-family:'JetBrains Mono',monospace; font-size:0.68rem;
                    font-weight:700; letter-spacing:0.1em; text-transform:uppercase;
                    color:{prof_color}; background:rgba(26,24,21,0.05);
                    padding:0.15rem 0.5rem; border-radius:2px;">PROFILO · {prof_name}</span>
            </div>
            <p style="font-size:0.9rem; color:#1A1815; margin:0; line-height:1.7;">{prof_text}</p>
        </div>""",
        unsafe_allow_html=True,
    )

    # ── Due colonne: concentrazione + geografia ───────────────────────────────
    fa_c1, fa_c2 = st.columns(2)

    with fa_c1:
        hhi_name, hhi_color, hhi_text = fa["hhi_label"]
        st.markdown(
            f"""<div class="glass-card" style="padding:1.1rem 1.4rem; height:100%;
                border-left:3px solid {hhi_color};">
                <div style="font-size:0.68rem; font-weight:700; letter-spacing:0.1em;
                    text-transform:uppercase; color:{hhi_color}; margin-bottom:0.4rem;">
                    CONCENTRAZIONE · {hhi_name}
                </div>
                <p style="font-size:0.86rem; color:#1A1815; margin:0 0 0.8rem; line-height:1.6;">{hhi_text}</p>
                <div style="font-size:0.8rem; color:#6B6258; margin-bottom:0.3rem;">
                    Top-5 posizioni: <span style="color:{hhi_color}; font-family:'JetBrains Mono',monospace;">
                    {fa['top5_pct']:.1f}%</span> del portafoglio
                </div>
                <div style="font-size:0.8rem; color:#6B6258;">{fa['sub_comment']}</div>
            </div>""",
            unsafe_allow_html=True,
        )

    with fa_c2:
        st.markdown(
            f"""<div class="glass-card" style="padding:1.1rem 1.4rem; height:100%;
                border-left:3px solid #1E3A8A;">
                <div style="font-size:0.68rem; font-weight:700; letter-spacing:0.1em;
                    text-transform:uppercase; color:#1E3A8A; margin-bottom:0.4rem;">
                    ESPOSIZIONE GEOGRAFICA
                </div>
                <p style="font-size:0.86rem; color:#1A1815; margin:0 0 0.8rem; line-height:1.6;">{fa['geo_note']}</p>
                {''.join(
                    f'<div style="display:flex; justify-content:space-between; align-items:center; '
                    f'margin-bottom:0.3rem;">'
                    f'<span style="font-size:0.8rem; color:#6B6258;">{g}</span>'
                    f'<span style="font-family:JetBrains Mono,monospace; font-size:0.8rem; color:#1E3A8A;">{p:.1f}%</span>'
                    f'</div>'
                    for g, p in fa['top_geo']
                )}
            </div>""",
            unsafe_allow_html=True,
        )

    st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)

    # ── Top 5 holdings table ──────────────────────────────────────────────────
    st.markdown(
        '<div style="font-size:0.68rem; font-weight:700; letter-spacing:0.1em; '
        'text-transform:uppercase; color:#A09A8E; margin-bottom:0.4rem;">'
        'TOP 5 POSIZIONI PER CONTROVALORE</div>',
        unsafe_allow_html=True,
    )
    fa_top5 = fa["top5"].copy()
    fa_top5["Controvalore"] = fa_top5["Controvalore"].apply(lambda x: f"€ {x:,.2f}" if pd.notna(x) else "N/D")
    fa_top5["Peso %"] = fa_top5["Peso %"].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "N/D")
    st.dataframe(fa_top5, use_container_width=True, hide_index=True)

    st.markdown("<div style='height:0.8rem'></div>", unsafe_allow_html=True)

    # ── Raccomandazioni ───────────────────────────────────────────────────────
    st.markdown(
        '<div style="font-size:0.68rem; font-weight:700; letter-spacing:0.1em; '
        'text-transform:uppercase; color:#A09A8E; margin-bottom:0.6rem;">'
        'RACCOMANDAZIONI OPERATIVE</div>',
        unsafe_allow_html=True,
    )
    for priority, rec_color, rec_text in fa["recommendations"]:
        st.markdown(
            f"""<div style="display:flex; gap:0.8rem; align-items:flex-start;
                margin-bottom:0.5rem; padding:0.7rem 1rem;
                background:#FBF7EE; border-radius:8px;
                border-left:2px solid {rec_color};">
                <span style="font-family:'JetBrains Mono',monospace; font-size:0.65rem;
                    font-weight:700; color:{rec_color}; white-space:nowrap;
                    padding-top:0.1rem;">{priority.upper()}</span>
                <span style="font-size:0.86rem; color:#1A1815; line-height:1.6;">{rec_text}</span>
            </div>""",
            unsafe_allow_html=True,
        )

# ── Data quality (in fondo alla pagina) ─────────────────────────────────────

st.markdown("<div style='height:2rem'></div>", unsafe_allow_html=True)
with st.expander(
    f"{'⚠️' if qr['missing_cells'] > 0 else '✅'}  Qualità dati — "
    f"{qr['rows']} righe · {qr['cols']} colonne · "
    f"{qr['missing_pct']}% valori mancanti",
    expanded=False,
):
    if qr["missing_cells"] == 0:
        st.markdown(
            '<p style="color:#3B5A36;">Nessun valore mancante — dataset completo.</p>',
            unsafe_allow_html=True,
        )
    else:
        miss_df = pd.DataFrame([
            {"Colonna": col, "Mancanti": v["count"], "%": v["pct"]}
            for col, v in qr["by_col"].items()
        ])
        st.dataframe(miss_df, use_container_width=True, hide_index=True)
    valid_ctv = df["Controvalore"].notna().sum() if "Controvalore" in df.columns else 0
    st.markdown(
        f'<p style="font-family:JetBrains Mono,monospace; font-size:0.75rem; color:#A09A8E;">'
        f'File: {", ".join(uf.name for uf in uploaded_files)} &nbsp;·&nbsp; {qr["rows"]} righe lette &nbsp;·&nbsp; '
        f'{len(df)} righe processate &nbsp;·&nbsp; {valid_ctv} con controvalore</p>',
        unsafe_allow_html=True,
    )
